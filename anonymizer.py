"""Anonymizer — bookkeeping-export anonymizer (building one increment at a time).

Subcommands correspond to CLAUDE.md build-order steps:
  roundtrip  (step 1) no-op round-trip gate: open the workbook, save it
             unchanged, prove dates/amounts/formulas/other tabs survive intact.
             GATES EVERYTHING; if it fails, stop and switch to the .txt fallback.
  inspect    (step 2) load both sheets + the config, print detected columns and
             per-sheet row counts, and summarize the three config lists.

Nothing past step 2 is implemented yet (no mapping, no replacement) — deliberate.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import random
import re
import secrets
import shutil
import sys
from pathlib import Path

import openpyxl

# Bumped on every commit going forward (also tagged in git as vN).
VERSION = "13"

# The three text columns rewritten by the regex pass (spec / CLAUDE.md).
TARGET_COLUMNS = ["Description", "Account", "Account #"]
# Institution: NOT faked, but normalized per account # so the trio stays
# consistent (this account/number -> one bank). Read + conditionally rewritten.
INSTITUTION_COLUMN = "Institution"
# The authoritative sheet: when an account #'s name/institution disagree across
# rows, the value seen on this sheet wins ("integrity dictated by transactions").
PRIMARY_SHEET = "Transactions"
# Config columns in the founder's combined file.
CONFIG_COLUMNS = ["Names", "Towns", "Blacklist"]
# Sheets the tool anonymizes. tiny.xlsx uses exactly these; the real workbook
# names its rule layer differently (RulesN) and will be mapped later.
DEFAULT_SHEETS = ["Transactions", "Roles"]

# Runs of 5+ digits inside descriptions (loan/membership/gov codes).
DIGIT_RUN_RE = re.compile(r"\d{5,}")

# Mixed letter+digit reference/confirmation codes (e.g. '6P038A5O05B',
# '99clacfxo', masked 'x0104'). A maximal alphanumeric run; a run counts as a
# code only if it has a letter and >=2 digits (so plain words and single-digit
# merchant tokens like '1STBANK' are left alone).
ALNUM_TOKEN_RE = re.compile(r"(?<![A-Za-z0-9])[A-Za-z0-9]{5,}(?![A-Za-z0-9])")


def is_code(token: str) -> bool:
    return any(c.isalpha() for c in token) and sum(c.isdigit() for c in token) >= 2


def extract_codes(text: str):
    return [t for t in ALNUM_TOKEN_RE.findall(text) if is_code(t)]


def fake_alnum(seed, code: str) -> str:
    """Deterministic same-shape fake for an alphanumeric code: each digit -> a
    random digit, each letter -> a random letter of the same case. Depends only
    on (seed, code), so it is identical wherever the code is seen."""
    r = random.Random(f"{seed}|alnum|{code}")
    out = []
    for c in code:
        if c.isdigit():
            out.append(str(r.randrange(10)))
        elif c.isupper():
            out.append(chr(65 + r.randrange(26)))
        elif c.islower():
            out.append(chr(97 + r.randrange(26)))
        else:
            out.append(c)
    return "".join(out)


# Word boundary that treats underscore as a separator, unlike \w (which
# includes '_'). Bookkeeping data glues two names together with an
# underscore (e.g. 'Ofir_Tali'); a plain \w boundary treats that as ONE
# word, so neither 'Ofir' nor 'Tali' would match inside it and both would
# survive un-anonymized. Real Unicode letters/digits still correctly block
# a partial match (e.g. 'Dom' won't wrongly match inside 'Domínguez').
WORDISH_PRE = r"(?<!(?!_)\w)"
WORDISH_POST = r"(?!(?!_)\w)"


def bare_token_re(token: str) -> "re.Pattern":
    """A whole digit token bounded by whitespace/punctuation on both sides
    (spec FR-ANON-08). Non-alphanumeric boundaries, so '0585' matches as a bare
    token but NOT inside a mixed code like 'x0585' or '1x1872TC' (those are
    handled by the alphanumeric-code rule instead)."""
    return re.compile(r"(?<![A-Za-z0-9])" + re.escape(token) + r"(?![A-Za-z0-9])")


def term_re(term: str) -> "re.Pattern":
    """Whole-word, case-insensitive match for a config name/town.

    Uses non-word lookarounds (equivalent to \\b for word-char-bounded terms,
    but safe for terms containing punctuation), so "Lee" is not found inside
    "sleeve". Case-insensitive because descriptions are often upper-cased while
    config entries are mixed-case (e.g. "Michael" vs "MICHAEL TEAM CK")."""
    return re.compile(r"(?<!\w)" + re.escape(term) + r"(?!\w)", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# Increment 1: no-op round-trip gate
# --------------------------------------------------------------------------- #

def round_trip(src: Path, dst: Path) -> None:
    """Open the workbook and save it unchanged to dst.

    Loads with data_only=False and keep_vba for .xlsm so formulas and macros
    are preserved. No cell is read, parsed, or modified.
    """
    keep_vba = src.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(src, data_only=False, keep_vba=keep_vba)
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()


_EMPTY = ("<empty>",)


def _cell_signature(cell):
    """What the round-trip must preserve.

    A cell with value None carries no data, so its internal type-tag and format
    are irrelevant — all empty cells collapse to one sentinel. (openpyxl re-tags
    empty inlineStr cells as generic empty on save; that is not data loss.)
    For cells that hold data, value + data_type + number_format must all match.
    """
    if cell.value is None:
        return _EMPTY
    return (cell.value, cell.data_type, cell.number_format)


def verify(src: Path, dst: Path):
    """Compare src and dst cell-by-cell across all sheets.

    Returns (ok, report_lines). A clean round-trip means: identical sheet list
    and identical (value, data_type, number_format) for every cell.
    """
    a = openpyxl.load_workbook(src, data_only=False)
    b = openpyxl.load_workbook(dst, data_only=False)
    lines = []
    ok = True

    if a.sheetnames != b.sheetnames:
        ok = False
        lines.append(f"  SHEET LIST differs: {a.sheetnames} != {b.sheetnames}")

    for sn in a.sheetnames:
        if sn not in b.sheetnames:
            continue
        wa, wb_ = a[sn], b[sn]
        diffs = 0
        first_diff = None
        max_r = max(wa.max_row or 0, wb_.max_row or 0)
        max_c = max(wa.max_column or 0, wb_.max_column or 0)
        ga = {(c.row, c.column): _cell_signature(c) for row in wa.iter_rows() for c in row}
        gb = {(c.row, c.column): _cell_signature(c) for row in wb_.iter_rows() for c in row}
        for k in sorted(set(ga) | set(gb)):
            if ga.get(k) != gb.get(k):
                diffs += 1
                if first_diff is None:
                    first_diff = (k, ga.get(k), gb.get(k))
        status = "OK" if diffs == 0 else f"{diffs} DIFF(S)"
        lines.append(f"  [{sn}] rows~{max_r} cols~{max_c}: {status}")
        if diffs:
            ok = False
            k, va, vb = first_diff
            lines.append(f"      first at cell (row={k[0]}, col={k[1]}): src={va!r} dst={vb!r}")

    a.close()
    b.close()
    return ok, lines


def cmd_roundtrip(args) -> int:
    src: Path = args.input
    if not src.exists():
        print(f"ERROR: input not found: {src}", file=sys.stderr)
        return 2
    dst: Path = args.out or (src.parent / "output" / f"{src.stem}_roundtrip{src.suffix}")

    print("Increment 1 - no-op round-trip")
    print(f"  source: {src}")
    print(f"  output: {dst}")
    round_trip(src, dst)

    ok, lines = verify(src, dst)
    print("Verification (value + data_type + number_format per cell):")
    for ln in lines:
        print(ln)
    print("RESULT:", "PASS - round-trip is clean; gate satisfied." if ok
          else "FAIL - round-trip NOT clean; per CLAUDE.md, stop and switch to .txt fallback.")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Increment 2: load workbook + config, report columns / counts
# --------------------------------------------------------------------------- #

def _clean(value) -> str:
    """Normalize a raw config cell to a stripped string ('' if empty)."""
    if value is None:
        return ""
    return str(value).strip()


def load_config(path: Path) -> dict:
    """Load the combined config (Names / Towns / Blacklist) from .csv or .xlsx.

    Returns {column: {"raw": [...], "distinct": [...]}} where 'raw' is every
    non-empty entry in file order and 'distinct' is deduplicated (first-seen
    order preserved). No mapping is built here — this is inspection only.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        # Excel saves "CSV UTF-8" and plain "CSV" differently (utf-8-sig vs
        # cp1252); a config hand-edited/re-saved in Excel can flip between
        # them, so fall back the same way bank CSVs already do.
        _enc, rows = read_csv_any(path)
    elif suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[_clean(c) for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raise ValueError(f"Unsupported config format: {suffix}")

    if not rows:
        raise ValueError(f"Config file is empty: {path}")

    header = [_clean(h) for h in rows[0]]
    index = {name: header.index(name) for name in CONFIG_COLUMNS if name in header}
    missing = [name for name in CONFIG_COLUMNS if name not in index]
    if missing:
        raise ValueError(
            f"Config is missing expected column(s) {missing}; found header {header}"
        )

    out = {}
    for name, col in index.items():
        raw = []
        for row in rows[1:]:
            val = _clean(row[col]) if col < len(row) else ""
            if val:
                raw.append(val)
        seen = set()
        distinct = []
        for v in raw:
            if v not in seen:
                seen.add(v)
                distinct.append(v)
        out[name] = {"raw": raw, "distinct": distinct}
    return out


def _header_and_count(ws):
    """Return (header_list, data_row_count) for a worksheet.

    Header is row 1; data row count is the number of subsequent rows that have
    at least one non-empty cell.
    """
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows, ()) or ())
    header = [_clean(h) for h in header]
    data = 0
    for r in rows:
        if any(c is not None and _clean(c) != "" for c in r):
            data += 1
    return header, data


def cmd_inspect(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    if not src.exists():
        print(f"ERROR: input workbook not found: {src}", file=sys.stderr)
        return 2
    if not cfg.exists():
        print(f"ERROR: config not found: {cfg}", file=sys.stderr)
        return 2

    print("Increment 2 - load workbook + config; report columns / counts")
    print(f"  workbook: {src}")
    print(f"  config:   {cfg}")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    wanted = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]

    print(f"\nWorkbook sheets present: {all_sheets}")
    missing_sheets = [s for s in (args.sheets or DEFAULT_SHEETS) if s not in all_sheets]
    if missing_sheets:
        print(f"  NOTE: expected sheet(s) not found: {missing_sheets}")

    ok = True
    for sn in wanted:
        ws = wb[sn]
        header, count = _header_and_count(ws)
        print(f"\n[{sn}]  rows(data)={count}  columns={len(header)}")
        print(f"  columns: {header}")
        found = {name: (header.index(name) + 1) for name in TARGET_COLUMNS if name in header}
        missing_cols = [name for name in TARGET_COLUMNS if name not in found]
        print(f"  target columns located (1-based col #): {found}")
        if missing_cols:
            ok = False
            print(f"  MISSING target column(s): {missing_cols}")
    wb.close()

    # Config summary
    try:
        config = load_config(cfg)
    except ValueError as e:
        print(f"\nERROR reading config: {e}", file=sys.stderr)
        return 1

    print("\nConfig lists:")
    for name in CONFIG_COLUMNS:
        raw = config[name]["raw"]
        distinct = config[name]["distinct"]
        dupes = len(raw) - len(distinct)
        extra = f"  ({dupes} duplicate(s) collapsed)" if dupes else ""
        print(f"  {name}: {len(raw)} entries, {len(distinct)} distinct{extra}")

    # Every Blacklist entry must be replaceable, i.e. present in the Names or
    # Towns lists (the spec says blacklist strings are copied into Names; town
    # blacklist entries are covered by the Towns list instead).
    replaceable = set(config["Names"]["distinct"]) | set(config["Towns"]["distinct"])
    bl_distinct = config["Blacklist"]["distinct"]
    covered = [b for b in bl_distinct if b in replaceable]
    print(f"  Blacklist entries covered by Names/Towns: {len(covered)}/{len(bl_distinct)}")
    not_covered = [b for b in bl_distinct if b not in replaceable]
    if not_covered:
        print(f"    NOTE: blacklist entries not in Names or Towns (would leak): {not_covered}")

    print("\nRESULT:", "PASS - workbook and config loaded; target columns present."
          if ok else "FAIL - a target column is missing on a sheet.")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Increment 3: extract identifiers, pooled across both sheets (detection only)
# --------------------------------------------------------------------------- #

def _target_indices(header):
    """0-based column indices of the three target columns in a header row."""
    return {name: header.index(name) for name in TARGET_COLUMNS if name in header}


def iter_target_cells(wb, sheets):
    """Yield (sheet, row_no, account, account_num, description, institution).

    row_no is 1-based worksheet row (header is row 1). Missing columns yield ""
    for that field. Values are cleaned strings.
    """
    for sn in sheets:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = [_clean(h) for h in (next(rows, ()) or ())]
        idx = {name: header.index(name) for name in
               (*TARGET_COLUMNS, INSTITUTION_COLUMN) if name in header}
        ai = idx.get("Account")
        ni = idx.get("Account #")
        di = idx.get("Description")
        ii = idx.get(INSTITUTION_COLUMN)
        for offset, r in enumerate(rows, start=2):
            def get(i):
                return _clean(r[i]) if i is not None and i < len(r) else ""
            acct, num, desc, inst = get(ai), get(ni), get(di), get(ii)
            if acct or num or desc or inst:
                yield sn, offset, acct, num, desc, inst


def last_four(account_num: str) -> str:
    """Trailing four digits of an account number, or '' if fewer than four."""
    digits = "".join(ch for ch in account_num if ch.isdigit())
    return digits[-4:] if len(digits) >= 4 else ""


def acct_key(account_num: str) -> str:
    """Normalized account identity: the last-four digits with leading zeros
    dropped ('xxxx0585', '0585', '585' -> '585'). Non-numeric numbers (e.g.
    'PP' for PayPal) key on the raw upper-cased string. '' if truly empty.

    This is THE account key: rules reference an account by its bare last-four
    (Excel drops the leading zeros), transactions by the masked 'xxxx####' form;
    both must resolve to the same identity so their fakes stay in sync."""
    digits = "".join(ch for ch in account_num if ch.isdigit())
    if digits:
        return str(int(digits[-4:]))
    return account_num.strip().upper()


def bare_forms(key: str) -> list:
    """The 4-digit form a numeric account key takes as a bare description token.

    Only the 4-digit (zero-padded) form is used for description matching and
    leak-scanning. The leading-zero-stripped int form (e.g. '225') is NOT — it
    is too short and collides with dollar amounts / quantities. The Account #
    COLUMN still normalizes every textual form via acct_key(), independently."""
    return [key.zfill(4)] if key.isdigit() else []


def canonical_value(pri_counter, all_counter) -> str:
    """Pick a canonical value: prefer the primary (Transactions) sheet's most
    frequent; fall back to the pooled most frequent. Ties -> alphabetical, so
    the choice is deterministic."""
    src = pri_counter if pri_counter else all_counter
    if not src:
        return ""
    top = max(src.values())
    return sorted(k for k, v in src.items() if v == top)[0]


def aggregate_identifiers(records):
    """Pool identifiers, keyed on the account NUMBER, from a stream of records.

    records: iterable of (is_primary, account, account_num, description,
    institution) tuples. is_primary marks the authoritative source (the
    Transactions table) so it can dictate canonical name/institution.

    Format-agnostic: fed by both the xlsx and the tab-delimited paths.

    Returns:
      accounts        : {number: {"names":Counter, "insts":Counter,
                                  "names_pri":Counter, "insts_pri":Counter}}
      name_only       : Counter of account names that appear with NO number
      real_account_names, real_account_numbers : sets (for the leakage scan)
      last_fours      : sorted distinct last-four strings
      digit_runs      : sorted distinct 5+ digit runs in descriptions
      cell_texts      : all non-empty target-cell strings (for locating config)
    """
    from collections import Counter

    accounts = {}          # acct_key -> aggregate
    name_only = Counter()
    names_with_number = set()
    real_names, real_nums = set(), set()
    digit_runs = set()
    alnum_codes = set()
    cell_texts = []

    for is_primary, acct, num, desc, inst in records:
        if acct:
            real_names.add(acct)
        if num:
            real_nums.add(num)
            key = acct_key(num)
            a = accounts.setdefault(key, {"names": Counter(), "insts": Counter(),
                                          "names_pri": Counter(), "insts_pri": Counter(),
                                          "forms": set()})
            a["forms"].add(num)
            if acct:
                a["names"][acct] += 1
                names_with_number.add(acct)
                if is_primary:
                    a["names_pri"][acct] += 1
            if inst:
                a["insts"][inst] += 1
                if is_primary:
                    a["insts_pri"][inst] += 1
        elif acct:
            name_only[acct] += 1
        for cell in (acct, num, desc):
            if cell:
                cell_texts.append(cell)
        if desc:
            digit_runs.update(DIGIT_RUN_RE.findall(desc))
            alnum_codes.update(extract_codes(desc))

    # A name is "name-only" only if it NEVER appears with a number.
    name_only = Counter({n: c for n, c in name_only.items() if n not in names_with_number})
    last_fours = sorted({lf for k in accounts for lf in bare_forms(k)})

    return {
        "accounts": accounts,
        "name_only": name_only,
        "real_account_names": real_names,
        "real_account_numbers": real_nums,
        "last_fours": last_fours,
        "digit_runs": sorted(digit_runs),
        "alnum_codes": sorted(alnum_codes),
        "cell_texts": cell_texts,
    }


def extract_identifiers(wb, sheets):
    """xlsx path: pool identifiers across sheets (Transactions is primary)."""
    return aggregate_identifiers(
        (sn == PRIMARY_SHEET, acct, num, desc, inst)
        for sn, _row, acct, num, desc, inst in iter_target_cells(wb, sheets))


def locate_terms(terms, cell_texts):
    """Return sorted (located, not_located) for config terms across pooled cells.

    A term is 'located' if it whole-word matches (case-insensitive) in any cell.
    """
    located, missing = [], []
    for term in terms:
        rx = term_re(term)
        if any(rx.search(t) for t in cell_texts):
            located.append(term)
        else:
            missing.append(term)
    return sorted(located), sorted(missing)


def cmd_extract(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    if not src.exists():
        print(f"ERROR: input workbook not found: {src}", file=sys.stderr)
        return 2
    if not cfg.exists():
        print(f"ERROR: config not found: {cfg}", file=sys.stderr)
        return 2

    print("Increment 3 - extract identifiers, pooled across both sheets")
    print(f"  workbook: {src}")
    print(f"  config:   {cfg}")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    sheets = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]
    print(f"  sheets pooled: {sheets}")

    ext = extract_identifiers(wb, sheets)
    wb.close()

    config = load_config(cfg)

    accts = ext["accounts"]
    print(f"\nAccounts (distinct by last-four key): {len(accts)}")
    for key in sorted(accts):
        info = accts[key]
        names = ", ".join(sorted(info["names"])) or "-"
        insts = sorted(info["insts"])
        forms = sorted(info["forms"])
        flag = "  <-- INSTITUTION CONFLICT" if len(insts) > 1 else ""
        print(f"  key={key!r:8} forms={forms}  names=[{names}]  institutions={insts}{flag}")
    if ext["name_only"]:
        print(f"\nAccount names with NO number ({len(ext['name_only'])}): {sorted(ext['name_only'])}")

    print(f"\n5+ digit runs in descriptions (distinct): {len(ext['digit_runs'])}")
    print(f"  {ext['digit_runs']}")

    names_located, names_missing = locate_terms(config["Names"]["distinct"], ext["cell_texts"])
    towns_located, towns_missing = locate_terms(config["Towns"]["distinct"], ext["cell_texts"])

    print(f"\nConfig Names located in workbook: {len(names_located)}/{len(config['Names']['distinct'])}")
    print(f"  located: {names_located}")
    print(f"\nConfig Towns located in workbook: {len(towns_located)}/{len(config['Towns']['distinct'])}")
    print(f"  located: {towns_located}")

    print("\nRESULT: PASS - identifiers extracted (detection only; no mapping built yet).")
    return 0


# --------------------------------------------------------------------------- #
# Increment 4: build ONE deterministic mapping, write the mapping table
# --------------------------------------------------------------------------- #

# Trailing account-type words preserved on fake account names so they read
# naturally (e.g. "Michael Team CK" -> "<fake person> CK").
ACCOUNT_TYPE_WORDS = {
    "CK", "CHK", "CHECKING", "CHECK", "SAVINGS", "SAV", "SV", "LLC", "INC",
    "CORP", "CO", "TRUST", "MGMT", "HOLDINGS", "OPS", "FUND", "ESCROW",
    # account-type / product roles (unambiguous banking terms, never surnames):
    "BUS", "BUSINESS", "BIZ", "VISA", "CARD", "CC", "LOAN", "PAYPAL", "CREDIT",
}

DATA_DIR = Path(__file__).resolve().parent / "data"
# Mapping table lives next to the code (script-relative, so it follows the
# folder if moved). It is the re-identification key - keep it private.
MAPPING_DIR = Path(__file__).resolve().parent / "Anon mapping"


def resolve_seed(args):
    """Return (seed, auto_generated).

    If --seed was given, use it. Otherwise mint a fresh random 6-hex-char seed
    (e.g. 'a7f3c2') that is recorded in the mapping and filenames, so the run
    stays reproducible afterward by passing that value back with --seed.
    """
    if args.seed:
        return args.seed, False
    return secrets.token_hex(3), True


def load_pools(data_dir: Path = DATA_DIR) -> dict:
    """Load the bundled fake-value pools (file order preserved = deterministic)."""
    def read(name):
        p = data_dir / name
        with open(p, encoding="utf-8") as f:
            items = [ln.strip() for ln in f if ln.strip()]
        if not items:
            raise ValueError(f"Pool file is empty: {p}")
        return items
    pools = {
        "first_names": read("first_names.txt"),
        "surnames": read("surnames.txt"),
        "towns": read("towns_pool.txt"),
    }
    streets = data_dir / "streets.txt"
    if streets.exists():                       # optional (Draft 9 address rule)
        pools["streets"] = [ln.strip() for ln in open(streets, encoding="utf-8") if ln.strip()]
    return pools


def _account_suffix(account_name: str) -> str:
    """Trailing account-type word if the real account name ends in one, else ''."""
    parts = account_name.split()
    if parts and parts[-1].upper().strip(".,") in ACCOUNT_TYPE_WORDS:
        return parts[-1]
    return ""


def build_mapping(identifiers: dict, config: dict, pools: dict, seed) -> dict:
    """Build ONE deterministic mapping from the pooled identifiers.

    Same seed -> identical mapping. Real values are consumed in a fixed order
    (sorted, category by category) and a single random.Random drives every pick;
    a fake that is already used, or that equals any real value, is redrawn.
    """
    rng = random.Random(seed)

    real_names = set(config["Names"]["distinct"])
    real_towns = set(config["Towns"]["distinct"])
    real_acct_names = set(identifiers["real_account_names"])
    real_acct_nums = set(identifiers["real_account_numbers"])
    real_last4 = set(identifiers["last_fours"])
    real_runs = set(identifiers["digit_runs"])

    used_names, used_towns, used_acct_names = set(), set(), set()
    used_last4, used_runs = set(), set()

    # A fake must not merely differ from real values — it must not CONTAIN a
    # real-side whole word either, or the leakage scan would flag it (e.g. the
    # pool surname "Michael" inside "John Michael"). Reject such candidates.
    forbidden_words = (set(config["Blacklist"]["distinct"]) | real_names
                       | real_towns | real_acct_names | real_acct_nums)
    fw = _alt(sorted(forbidden_words))
    forbid_rx = re.compile(rf"(?<!\w){fw}(?!\w)", re.IGNORECASE) if fw else None

    def contains_forbidden(text: str) -> bool:
        return bool(forbid_rx and forbid_rx.search(text))

    def draw(make, used, forbidden, what, reject=None):
        for _ in range(10000):
            cand = make()
            if cand in used or cand in forbidden:
                continue
            if reject and reject(cand):
                continue
            used.add(cand)
            return cand
        raise RuntimeError(f"Could not find a unique fake {what}; pool too small?")

    def fake_person():
        return f"{rng.choice(pools['first_names'])} {rng.choice(pools['surnames'])}"

    def fake_single():
        return rng.choice(pools["first_names"])

    def fake_four():
        return f"{rng.randrange(10000):04d}"

    # Config Names (sorted) -> fake names matching the real word count: a
    # one-word real ("Michael", "SBA") maps to a single fake name; a two-or-
    # more-word real ("Michael Barrios", "42 Van Winkle") maps to two words.
    names = {}
    for real in sorted(real_names):
        make = fake_single if len(real.split()) <= 1 else fake_person
        names[real] = draw(make, used_names, real_names | real_acct_names,
                           "name", reject=contains_forbidden)

    # Config Towns (sorted) -> fake towns.
    towns = {}
    for real in sorted(real_towns):
        towns[real] = draw(lambda: rng.choice(pools["towns"]), used_towns, real_towns,
                           "town", reject=contains_forbidden)

    # Account NAMES map to fake names keyed on the NAME (one fake per distinct
    # real name), so a name shared by two accounts stays consistent everywhere.
    all_names = set(identifiers["real_account_names"]) | set(identifiers["name_only"])
    account_names = {}
    for real in sorted(all_names):
        suffix = _account_suffix(real)

        def make_acct(suffix=suffix):
            person = fake_person()
            return f"{person} {suffix}".strip() if suffix else person
        account_names[real] = draw(make_acct, used_acct_names, real_acct_names,
                                   "account name", reject=contains_forbidden)

    # Account NUMBERS, keyed on the normalized last-four. One fake last-four per
    # account; institution is dictated by the Transactions sheet.
    accounts = []
    last_four_map = {}
    acct_info = identifiers["accounts"]
    for key in sorted(acct_info):
        info = acct_info[key]
        canon_inst = canonical_value(info["insts_pri"], info["insts"])
        flf = draw(fake_four, used_last4, real_last4, "last-four")
        # The account's 4-digit last-four maps to the fake last-four for bare
        # description tokens (transfer references).
        for form in bare_forms(key):
            last_four_map[form] = flf

        accounts.append({
            "key": key,
            "real_forms": sorted(info["forms"]),
            "real_names": sorted(info["names"]),
            "fake_last4": flf,
            "fake_number": "xxxx" + flf,
            "canonical_institution": canon_inst,
            "real_institutions": sorted(info["insts"]),
            "institution_conflict": len(info["insts"]) > 1,
        })

    # 5+ digit runs (sorted) -> fake runs of the same length.
    digit_runs = {}
    for real in identifiers["digit_runs"]:
        n = len(real)
        def make_run(n=n):
            return "".join(str(rng.randrange(10)) for _ in range(n))
        digit_runs[real] = draw(make_run, used_runs, real_runs, "digit run")

    # Mixed letter+digit codes -> deterministic same-shape fakes (per value).
    alnum_codes = {code: fake_alnum(seed, code) for code in identifiers.get("alnum_codes", [])}

    return {
        "seed": seed,
        "names": names,
        "towns": towns,
        "account_names": account_names,
        "accounts": accounts,
        "last_four": last_four_map,
        "digit_runs": digit_runs,
        "alnum_codes": alnum_codes,
    }


def _swap_last_four(real_num: str, fake4: str) -> str:
    """Replace the last four digits of real_num with fake4, keeping the rest.

    Preserves masking prefixes like 'xxxx' so 'xxxx2077' -> 'xxxx' + fake4.
    """
    digits_seen = 0
    chars = list(real_num)
    for i in range(len(chars) - 1, -1, -1):
        if chars[i].isdigit():
            chars[i] = fake4[3 - digits_seen]
            digits_seen += 1
            if digits_seen == 4:
                break
    return "".join(chars)


def write_mapping(mapping: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False, sort_keys=True)
        f.write("\n")


def cmd_map(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    if not src.exists():
        print(f"ERROR: input workbook not found: {src}", file=sys.stderr)
        return 2
    if not cfg.exists():
        print(f"ERROR: config not found: {cfg}", file=sys.stderr)
        return 2

    args.seed, auto = resolve_seed(args)
    print("Increment 4 - build ONE deterministic mapping; write mapping table")
    print(f"  workbook: {src}")
    print(f"  config:   {cfg}")
    print(f"  seed:     {args.seed}" + ("   (auto-generated)" if auto else ""))

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    sheets = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]
    identifiers = extract_identifiers(wb, sheets)
    wb.close()
    config = load_config(cfg)
    pools = load_pools()

    mapping = build_mapping(identifiers, config, pools, args.seed)

    out = args.map_out or (MAPPING_DIR / f"mapping.{args.seed}.json")
    write_mapping(mapping, out)

    conflicts = [a for a in mapping["accounts"] if a["institution_conflict"]]
    print(f"\nMapping written (re-identification key - keep private): {out}")
    print(f"  config names mapped:   {len(mapping['names'])}")
    print(f"  config towns mapped:   {len(mapping['towns'])}")
    print(f"  account names mapped:  {len(mapping['account_names'])}")
    print(f"  accounts (by last-4):  {len(mapping['accounts'])}")
    print(f"  last-four forms mapped:{len(mapping['last_four'])}")
    print(f"  5+ digit runs:         {len(mapping['digit_runs'])}")
    print(f"  accounts w/ institution conflict (normalized): {len(conflicts)}")

    print("\nSample account mappings (key / real forms / institution -> fake):")
    for acc in mapping["accounts"][:6]:
        flag = "  [normalized: " + ", ".join(acc["real_institutions"]) + "]" if acc["institution_conflict"] else ""
        print(f"  key {acc['key']!r} forms={acc['real_forms']} inst={acc['canonical_institution']!r} "
              f"-> {acc['fake_number']!r}, kept {acc['canonical_institution']!r}{flag}")

    print("\nRESULT: PASS - mapping built and written (deterministic; not applied yet).")
    return 0


# --------------------------------------------------------------------------- #
# Increment 5: single-pass replacement over the three columns of both sheets
# --------------------------------------------------------------------------- #

def _alt(literals):
    """Alternation of escaped literals, longest first (so 'Michael Barrios'
    is tried before 'Michael'). Returns None if there are no literals."""
    uniq = sorted(set(literals), key=lambda s: (-len(s), s))
    if not uniq:
        return None
    return "(?:" + "|".join(re.escape(s) for s in uniq) + ")"


def build_replacer(mapping: dict):
    """Compile ONE combined pattern for all rules and return (pattern, repl, stats).

    A single re.sub with this pattern is the whole replacement pass: the regex
    engine scans left-to-right and never re-examines substituted text, giving
    single-pass / no-chaining for free. Alternatives are ordered most-specific
    first: account name, config name, config town, account number, 5+ digit run,
    bare last-four. Names/towns/accounts match case-insensitively on whole-word
    boundaries; the bare last-four uses digit boundaries (?<!\\d)NNNN(?!\\d).
    """
    accounts = mapping["accounts"]
    # Account NAMES are keyed on the name (one fake per distinct real name).
    acctname_map = {r.lower(): f for r, f in mapping["account_names"].items()}
    # Masked account-number forms ('xxxx0585') that may appear in free text map
    # to the account's 'xxxx'+fake last-four. (The Account # column itself is
    # normalized directly in apply, not via this regex.)
    acctnum_map = {}
    for a in accounts:
        for form in a["real_forms"]:
            if any(not ch.isdigit() for ch in form):   # has 'xxxx'-style prefix
                acctnum_map[form.lower()] = a["fake_number"]
    name_map = {r.lower(): f for r, f in mapping["names"].items()}
    town_map = {r.lower(): f for r, f in mapping["towns"].items()}
    run_map = dict(mapping["digit_runs"])
    last4_map = dict(mapping["last_four"])
    alnum_map = dict(mapping.get("alnum_codes", {}))
    seed = mapping.get("seed", "")

    parts = []
    a = _alt(list(acctname_map))
    if a:
        parts.append(rf"(?P<acctname>(?<!\w){a}(?!\w))")
    n = _alt(list(mapping["names"]))
    if n:
        parts.append(rf"(?P<name>(?<!\w){n}(?!\w))")
    t = _alt(list(mapping["towns"]))
    if t:
        parts.append(rf"(?P<town>(?<!\w){t}(?!\w))")
    an = _alt(list(acctnum_map))
    if an:
        parts.append(rf"(?P<acctnum>(?<!\w){an}(?!\w))")
    parts.append(r"(?P<run>(?<![A-Za-z0-9])\d{5,}(?![A-Za-z0-9]))")
    if last4_map:
        l4 = "(?:" + "|".join(re.escape(k) for k in sorted(last4_map)) + ")"
        parts.append(rf"(?P<last4>(?<![A-Za-z0-9]){l4}(?![A-Za-z0-9]))")
    # Alphanumeric codes last: only fires where nothing more specific matched.
    parts.append(r"(?P<alcode>(?<![A-Za-z0-9])[A-Za-z0-9]{5,}(?![A-Za-z0-9]))")

    pattern = re.compile("|".join(parts), re.IGNORECASE)

    stats = {}

    def repl(m):
        kind = m.lastgroup
        text = m.group()
        if kind == "acctname":
            out = acctname_map[text.lower()]
        elif kind == "name":
            out = name_map[text.lower()]
        elif kind == "town":
            out = town_map[text.lower()]
        elif kind == "acctnum":
            out = acctnum_map[text.lower()]
        elif kind == "run":
            out = run_map.get(text, text)
            if text not in run_map:
                kind = "run_unmapped"
        elif kind == "last4":
            out = last4_map.get(text, text)
        elif kind == "alcode":
            if is_code(text):
                out = alnum_map.get(text) or fake_alnum(seed, text)
            else:
                out = text            # plain word / single-digit token: leave it
                kind = "alcode_skip"
        else:
            out = text
        stats[kind] = stats.get(kind, 0) + 1
        return out

    return pattern, repl, stats


def account_output_maps(mapping):
    """Return (number_by_key, inst_by_key):
      number_by_key : acct_key -> 'xxxx'+fake last-four (Account # column output)
      inst_by_key   : acct_key -> canonical real institution (when it has one)."""
    number_by_key = {a["key"]: a["fake_number"] for a in mapping["accounts"]}
    inst_by_key = {a["key"]: a["canonical_institution"]
                   for a in mapping["accounts"] if a["canonical_institution"]}
    return number_by_key, inst_by_key


def apply_to_workbook(wb, sheets, pattern, repl, number_by_key=None, inst_by_key=None):
    """Anonymize the target columns and normalize the account-number/institution.

    Account # column -> normalized to 'xxxx'+fake last-four (keyed on last-four).
    Account name + Description -> regex pass. Institution -> canonical per key.
    Every other cell is left exactly as opened. Returns counts dict.
    """
    number_by_key = number_by_key or {}
    inst_by_key = inst_by_key or {}
    changed = insts = 0
    for sn in sheets:
        ws = wb[sn]
        header = [_clean(c.value) for c in ws[1]]
        col = {name: header.index(name) + 1 for name in
               (*TARGET_COLUMNS, INSTITUTION_COLUMN) if name in header}
        num_col = col.get("Account #")
        inst_col = col.get(INSTITUTION_COLUMN)
        regex_cols = {col[n] for n in ("Description", "Account") if n in col}
        for row in ws.iter_rows(min_row=2):
            cells = {c.column: c for c in row}
            real_num = _clean(cells[num_col].value) if (num_col in cells and cells.get(num_col) and cells[num_col].value is not None) else ""
            key = acct_key(real_num) if real_num else ""
            # Account name + Description via regex.
            for cidx in regex_cols:
                cell = cells.get(cidx)
                if cell is not None and cell.value is not None:
                    original = str(cell.value)
                    new = pattern.sub(repl, original)
                    if new != original:
                        cell.value = new
                        changed += 1
            # Account # column -> normalized fake.
            if num_col and key in number_by_key:
                cell = cells.get(num_col)
                if cell is not None and str(cell.value) != number_by_key[key]:
                    cell.value = number_by_key[key]
                    changed += 1
            # Institution -> canonical for this account.
            if inst_col and key in inst_by_key:
                cell = cells.get(inst_col)
                if cell is not None and inst_by_key[key] and _clean(cell.value) != inst_by_key[key]:
                    cell.value = inst_by_key[key]
                    insts += 1
    return {"cells": changed, "institutions": insts}


def cmd_apply(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    if not src.exists():
        print(f"ERROR: input workbook not found: {src}", file=sys.stderr)
        return 2
    if not cfg.exists():
        print(f"ERROR: config not found: {cfg}", file=sys.stderr)
        return 2

    args.seed, auto = resolve_seed(args)
    print("Increment 5 - single-pass replacement over three columns, both sheets")
    print(f"  workbook: {src}")
    print(f"  config:   {cfg}")
    print(f"  seed:     {args.seed}" + ("   (auto-generated)" if auto else ""))

    # Build the ONE mapping (same as increment 4).
    rwb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = rwb.sheetnames
    sheets = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]
    identifiers = extract_identifiers(rwb, sheets)
    rwb.close()
    config = load_config(cfg)
    pools = load_pools()
    mapping = build_mapping(identifiers, config, pools, args.seed)

    pattern, repl, stats = build_replacer(mapping)

    # Load a writable copy, edit target columns + institution, save parallel.
    keep_vba = src.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(src, data_only=False, keep_vba=keep_vba)
    num_by_key, inst_by_key = account_output_maps(mapping)
    counts = apply_to_workbook(wb, sheets, pattern, repl, num_by_key, inst_by_key)
    dst = args.out or (src.parent / "output" / f"{src.stem}_anon.{args.seed}{src.suffix}")
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()

    print(f"\n  sheets processed: {sheets}")
    print(f"  cells changed:    {counts['cells']}")
    print(f"  institutions normalized: {counts['institutions']}")
    print("  replacements by rule:")
    for kind in ("acctname", "name", "town", "acctnum", "run", "last4", "run_unmapped"):
        if kind in stats:
            print(f"    {kind:12s}: {stats[kind]}")
    print(f"\n  anonymized workbook written: {dst}")

    # Small before/after sample from the Description column of each sheet.
    orig = openpyxl.load_workbook(src, read_only=True, data_only=True)
    anon = openpyxl.load_workbook(dst, read_only=True, data_only=True)
    print("\nSample Description before -> after:")
    for sn in sheets:
        ho = [_clean(h) for h in next(orig[sn].iter_rows(values_only=True), ())]
        di = ho.index("Description") if "Description" in ho else None
        if di is None:
            continue
        orows = list(orig[sn].iter_rows(min_row=2, values_only=True))
        arows = list(anon[sn].iter_rows(min_row=2, values_only=True))
        print(f"  [{sn}]")
        for o, a in list(zip(orows, arows))[:6]:
            print(f"    - {o[di]!r}")
            print(f"      {a[di]!r}")
    orig.close()
    anon.close()

    ok = "run_unmapped" not in stats
    print("\nRESULT:", "PASS - replacement pass complete."
          if ok else "WARN - a 5+ digit run was not in the mapping (see run_unmapped).")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Increment 6: blacklist / real-side leakage scan (both sheets)
# --------------------------------------------------------------------------- #

def build_leak_scanners(identifiers: dict, config: dict):
    """Compile matchers for everything that must NOT survive in the output.

    Two patterns: word-bounded (names, towns, account names, blacklist words,
    account numbers) and digit-bounded (real last-fours and 5+ digit runs).
    Returns (word_rx, digit_rx); either may be None if its set is empty.
    """
    word_terms = set(config["Blacklist"]["distinct"])
    word_terms |= set(config["Names"]["distinct"])
    word_terms |= set(config["Towns"]["distinct"])
    word_terms |= set(identifiers["real_account_names"])
    word_terms |= set(identifiers["real_account_numbers"])

    digit_terms = set(identifiers["last_fours"]) | set(identifiers["digit_runs"])

    w = _alt(sorted(word_terms))
    word_rx = re.compile(rf"(?<!\w){w}(?!\w)", re.IGNORECASE) if w else None
    digit_rx = None
    if digit_terms:
        d = "(?:" + "|".join(re.escape(t) for t in sorted(digit_terms)) + ")"
        digit_rx = re.compile(rf"(?<![A-Za-z0-9]){d}(?![A-Za-z0-9])")
    return word_rx, digit_rx


# Columns that can carry anonymized identifiers, so are worth leak-scanning.
# Amount/Date/Balance/Seq are numbers; Institution is intentionally kept real.
SCAN_COLUMNS = {"description", "account", "account #"}


def scan_workbook(path: Path, sheets, word_rx, digit_rx):
    """Scan the identifier-bearing columns (SCAN_COLUMNS) on the given sheets.

    Skips numeric columns (Amount/Date/Balance) and the kept-real Institution
    column, whose real values would otherwise register as false leaks.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    leaks = []
    for sn in sheets:
        if sn not in wb.sheetnames:
            continue
        it = wb[sn].iter_rows(values_only=True)
        header = [_clean(h).lower() for h in (next(it, ()) or ())]
        cols = {i for i, h in enumerate(header) if h in SCAN_COLUMNS}
        for r, row in enumerate(it, start=2):
            for c in cols:
                val = row[c] if c < len(row) else None
                if val is None:
                    continue
                s = str(val)
                for rx in (word_rx, digit_rx):
                    if rx and rx.search(s):
                        leaks.append((sn, r, c + 1, s, rx.search(s).group()))
                        break
    wb.close()
    return leaks


def leakage_check(identifiers, config, src: Path, anon: Path, sheets):
    """Return (ok, before_leaks, after_leaks).

    ok requires zero leaks in the output AND a non-empty self-test on the
    original (so a broken matcher can't masquerade as a clean run).
    """
    word_rx, digit_rx = build_leak_scanners(identifiers, config)
    before = scan_workbook(src, sheets, word_rx, digit_rx)
    after = scan_workbook(anon, sheets, word_rx, digit_rx)
    ok = len(after) == 0 and len(before) > 0
    return ok, before, after


def cmd_scan(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    anon: Path = args.anon or (src.parent / "output" / f"{src.stem}_anon{src.suffix}")
    for label, path in [("input", src), ("config", cfg), ("anonymized", anon)]:
        if not path.exists():
            print(f"ERROR: {label} not found: {path}", file=sys.stderr)
            return 2

    print("Increment 6 - blacklist / real-side leakage scan")
    print(f"  original:   {src}")
    print(f"  anonymized: {anon}")

    rwb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = rwb.sheetnames
    sheets = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]
    identifiers = extract_identifiers(rwb, sheets)
    rwb.close()
    config = load_config(cfg)

    ok, before, leaks = leakage_check(identifiers, config, src, anon, sheets)
    print(f"\nSelf-test - leaks found in ORIGINAL (should be > 0): {len(before)}")
    if not before:
        print("  WARNING: scanner found nothing in the original; matchers may be broken.")
    print(f"Leaks found in ANONYMIZED output (must be 0): {len(leaks)}")
    for sn, r, c, s, hit in leaks[:25]:
        print(f"  LEAK {sn}!R{r}C{c}: matched {hit!r} in {s!r}")
    if len(leaks) > 25:
        print(f"  ... and {len(leaks) - 25} more")

    print("\nRESULT:", "PASS - no blacklist/real-side string survives."
          if ok else "FAIL - leakage detected (or scanner self-test failed).")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Increment 7: hit-count check (account/transfer rules fire the same # of times)
# --------------------------------------------------------------------------- #

def count_all(path: Path, sheets, matchers):
    """One pass over a workbook; count matches per (key, rx, columns) matcher.

    columns is a set of column names to restrict counting to. Returns {key: n}.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    counts = {key: 0 for key, _rx, _cols in matchers}
    for sn in sheets:
        if sn not in wb.sheetnames:
            continue
        it = wb[sn].iter_rows(values_only=True)
        header = [_clean(h) for h in (next(it, ()) or ())]
        colidx = {name: i for i, name in enumerate(header)}
        for row in it:
            for key, rx, cols in matchers:
                for cn in cols:
                    i = colidx.get(cn)
                    if i is not None and i < len(row) and row[i] is not None:
                        counts[key] += len(rx.findall(str(row[i])))
    wb.close()
    return counts


def hitcount_matchers(mapping):
    """Return [(key, real_rx, real_cols, fake_rx, fake_cols, mode)] for the
    hit-count check, shared by the xlsx and TSV paths. `mode` controls gating:
      'exact' - before must equal after (fail on any difference)
      'drop'  - fail only if after < before (a genuinely missed conversion)
      'info'  - reported, never fails

    Account names are 'info': they overlap each other and config names, so exact
    counts are ill-defined (survival is guaranteed by the leakage scan). Bare
    last-fours are 'drop': a drop means a transfer reference was missed; a jump
    is a benign coincidence with an innocent number. 5+ digit runs are 'exact'
    (unique, collision-free)."""
    TARGET = set(TARGET_COLUMNS)
    DESC = {"Description"}
    gated = []
    for real, fake in mapping["account_names"].items():
        gated.append((f"acctname:{real}", term_re(real), TARGET, term_re(fake), TARGET, "info"))
    for real_bare, fake_bare in mapping["last_four"].items():
        gated.append((f"last4:{real_bare}", bare_token_re(real_bare), DESC,
                      bare_token_re(fake_bare), DESC, "drop"))
    for rr, fr in mapping["digit_runs"].items():
        gated.append((f"run:{rr}", bare_token_re(rr), DESC, bare_token_re(fr), DESC, "exact"))
    return gated


def hitcount_failed(mode, before, after):
    """True if this matcher's before/after violates its gate mode."""
    if mode == "exact":
        return before != after
    if mode == "drop":
        return after < before
    return False  # info


def hitcount_check(mapping, src: Path, anon: Path, sheets):
    """xlsx: compare real-count-before vs fake-count-after for gated identifiers.
    Returns (ok, rows) where rows is [(key, before, after)]."""
    gated = hitcount_matchers(mapping)
    before = count_all(src, sheets, [(k, rx, c) for k, rx, c, _f, _fc, _m in gated])
    after = count_all(anon, sheets, [(k, rx, c) for k, _r, _rc, rx, c, _m in gated])
    rows = [(k, before[k], after[k], m) for k, *_rest, m in gated]
    ok = not any(hitcount_failed(m, b, a) for _k, b, a, m in rows)
    return ok, rows


def cmd_hitcount(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    anon: Path = args.anon or (src.parent / "output" / f"{src.stem}_anon.{args.seed}{src.suffix}")
    for label, path in [("input", src), ("config", cfg), ("anonymized", anon)]:
        if not path.exists():
            print(f"ERROR: {label} not found: {path}", file=sys.stderr)
            return 2

    print("Increment 7 - hit-count check (account/transfer rules)")
    print(f"  seed: {args.seed}")

    rwb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = rwb.sheetnames
    sheets = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]
    identifiers = extract_identifiers(rwb, sheets)
    rwb.close()
    config = load_config(cfg)
    pools = load_pools()
    mapping = build_mapping(identifiers, config, pools, args.seed)

    ok, rows = hitcount_check(mapping, src, anon, sheets)

    print("\n  rule                              before  after")
    for k, b, a, m in rows:
        flag = "  <-- FAIL" if hitcount_failed(m, b, a) else (
            "" if b == a else ("  (drop)" if a < b else "  (jump/benign)"))
        print(f"  {k:34s}  {b:5d}  {a:5d}{flag}")

    TARGET = set(TARGET_COLUMNS)

    # Informational: names/towns totals (overlapping sub-tokens make per-term
    # equality ill-defined, so these are reported, not gated).
    name_real = [(f"n{i}", term_re(r), TARGET) for i, r in enumerate(mapping["names"])]
    name_fake = [(f"n{i}", term_re(f), TARGET) for i, f in enumerate(mapping["names"].values())]
    town_real = [(f"t{i}", term_re(r), TARGET) for i, r in enumerate(mapping["towns"])]
    town_fake = [(f"t{i}", term_re(f), TARGET) for i, f in enumerate(mapping["towns"].values())]
    nb = sum(count_all(src, sheets, name_real).values())
    na = sum(count_all(anon, sheets, name_fake).values())
    tb = sum(count_all(src, sheets, town_real).values())
    ta = sum(count_all(anon, sheets, town_fake).values())
    print(f"\n  (info) config names: before={nb} after={na}   towns: before={tb} after={ta}")
    print("  (names/towns are informational; sub-token overlap is handled by the account rule)")

    print("\nRESULT:", "PASS - all account/transfer hit-counts hold."
          if ok else "FAIL - a hit-count changed (missed or over-replacement).")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Increment 8: one-command runner + consolidated report
# --------------------------------------------------------------------------- #

def produce_anon(src: Path, dst: Path, mapping: dict, sheets) -> dict:
    """Apply the mapping to a fresh copy of src, save to dst. Returns counts."""
    pattern, repl, _stats = build_replacer(mapping)
    keep_vba = src.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(src, data_only=False, keep_vba=keep_vba)
    num_by_key, inst_by_key = account_output_maps(mapping)
    counts = apply_to_workbook(wb, sheets, pattern, repl, num_by_key, inst_by_key)
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()
    return counts


def verify_nontarget(src: Path, dst: Path, sheets):
    """Everything EXCEPT the intentionally-edited columns must be byte-intact.

    Compares source vs output cell-by-cell (value+type+format, empty-cell aware):
    on the anonymized sheets the three target columns AND Institution (which is
    normalized) are skipped; all other columns AND all other sheets must match
    exactly. Returns (ok, lines).
    """
    a = openpyxl.load_workbook(src, data_only=False)
    b = openpyxl.load_workbook(dst, data_only=False)
    lines, ok = [], True
    if a.sheetnames != b.sheetnames:
        ok = False
        lines.append(f"  SHEET LIST differs: {a.sheetnames} != {b.sheetnames}")
    editable = (*TARGET_COLUMNS, INSTITUTION_COLUMN)
    for sn in a.sheetnames:
        if sn not in b.sheetnames:
            continue
        wa, wb_ = a[sn], b[sn]
        skip = set()
        if sn in sheets:
            header = [_clean(c.value) for c in wa[1]]
            skip = {header.index(n) + 1 for n in editable if n in header}
        ga = {(c.row, c.column): _cell_signature(c)
              for row in wa.iter_rows() for c in row if c.column not in skip}
        gb = {(c.row, c.column): _cell_signature(c)
              for row in wb_.iter_rows() for c in row if c.column not in skip}
        diffs = sum(1 for k in set(ga) | set(gb) if ga.get(k) != gb.get(k))
        if diffs:
            ok = False
            lines.append(f"  [{sn}] {diffs} non-target cell(s) changed")
    a.close()
    b.close()
    return ok, lines


def content_identical(p1: Path, p2: Path):
    """True if two workbooks have identical cell content across all sheets."""
    a = openpyxl.load_workbook(p1, data_only=False)
    b = openpyxl.load_workbook(p2, data_only=False)
    same = a.sheetnames == b.sheetnames
    if same:
        for sn in a.sheetnames:
            ga = {(c.row, c.column): _cell_signature(c) for row in a[sn].iter_rows() for c in row}
            gb = {(c.row, c.column): _cell_signature(c) for row in b[sn].iter_rows() for c in row}
            if ga != gb:
                same = False
                break
    a.close()
    b.close()
    return same


def cmd_run(args) -> int:
    src: Path = args.input
    cfg: Path = args.config
    if not src.exists():
        print(f"ERROR: input not found: {src}", file=sys.stderr)
        return 2
    if not cfg.exists():
        print(f"ERROR: config not found: {cfg}", file=sys.stderr)
        return 2

    args.seed, auto = resolve_seed(args)
    out = args.out or (src.parent / "output" / f"{src.stem}_anon.{args.seed}{src.suffix}")
    map_out = args.map_out or (MAPPING_DIR / f"mapping.{args.seed}.json")

    print("=" * 64)
    print("ANONYMIZER RUN")
    print("=" * 64)
    print(f"  input:   {src}")
    print(f"  config:  {cfg}")
    print(f"  seed:    {args.seed}" + ("   (auto-generated - pass this to --seed to reproduce)" if auto else ""))

    rwb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    all_sheets = rwb.sheetnames
    sheets = args.sheets or [s for s in DEFAULT_SHEETS if s in all_sheets]
    missing = [s for s in (args.sheets or DEFAULT_SHEETS) if s not in all_sheets]
    identifiers = extract_identifiers(rwb, sheets)
    rwb.close()
    config = load_config(cfg)
    pools = load_pools()
    print(f"  sheets:  {sheets}" + (f"   (NOTE missing: {missing})" if missing else ""))

    # Build + write the mapping, produce the anonymized workbook.
    mapping = build_mapping(identifiers, config, pools, args.seed)
    write_mapping(mapping, map_out)
    counts = produce_anon(src, out, mapping, sheets)
    print(f"  output:  {out}")
    print(f"  mapping: {map_out}  (re-identification key - keep private)")
    print(f"  cells changed: {counts['cells']}   institutions normalized: {counts['institutions']}")

    # Checks.
    results = []

    int_ok, int_lines = verify_nontarget(src, out, sheets)
    results.append(("Non-target integrity (dates/amounts/formulas/other tabs)", int_ok))
    for ln in int_lines:
        print(ln)

    leak_ok, before, leaks = leakage_check(identifiers, config, src, out, sheets)
    results.append((f"Leakage scan (orig leaks={len(before)} -> output leaks={len(leaks)})", leak_ok))
    for sn, r, c, s, hit in leaks[:10]:
        print(f"    LEAK {sn}!R{r}C{c}: {hit!r} in {s!r}")

    hc_ok, rows = hitcount_check(mapping, src, out, sheets)
    bad = [(k, b, a) for k, b, a, m in rows if hitcount_failed(m, b, a)]
    results.append((f"Hit-count check ({len(rows)} rules, {len(bad)} failing)", hc_ok))
    for k, b, a in bad[:10]:
        print(f"    MISMATCH {k}: before={b} after={a}")

    # Determinism: a second full run to a temp file must be content-identical.
    tmp = out.parent / f".{out.stem}.det{out.suffix}"
    mapping2 = build_mapping(identifiers, config, pools, args.seed)
    produce_anon(src, tmp, mapping2, sheets)
    det_ok = content_identical(out, tmp)
    try:
        tmp.unlink()
    except OSError:
        pass
    results.append(("Determinism (same seed -> identical content)", det_ok))

    print("\n" + "-" * 64)
    print("REPORT")
    print("-" * 64)
    all_ok = True
    for label, passed in results:
        print(f"  [{'PASS' if passed else 'FAIL'}]  {label}")
        all_ok = all_ok and passed
    print("-" * 64)
    print("OVERALL:", "PASS - anonymized workbook is safe to share." if all_ok
          else "FAIL - do NOT share; see failures above.")
    print("=" * 64)
    return 0 if all_ok else 1


# --------------------------------------------------------------------------- #
# Tab-delimited (TSV) path — the robust fallback for complex .xlsm workbooks.
# Reads Excel's "Text (Tab delimited)" exports (Windows-1252), never opens the
# workbook, and writes flat TSV output. Only the 7 important columns are kept.
# --------------------------------------------------------------------------- #

TSV_ENCODING = "cp1252"          # Excel "Text (Tab delimited)" on Windows
TSV_DELIM = "\t"
# Columns kept in the output, in this order. Everything else is dropped (which
# also means non-target columns like Notes / Full Description can't leak).
IMPORTANT_COLUMNS = ["Seq", "Date", "Description", "Amount",
                     "Account", "Account #", "Institution"]


def read_tsv(path: Path):
    """Return (header, rows) of raw strings. Fields are never parsed."""
    with open(path, encoding=TSV_ENCODING, newline="") as f:
        data = list(csv.reader(f, delimiter=TSV_DELIM))
    return (data[0], data[1:]) if data else ([], [])


def write_tsv(path: Path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding=TSV_ENCODING, newline="") as f:
        csv.writer(f, delimiter=TSV_DELIM).writerows(rows)


def _ci_index(header, name):
    """Case-insensitive index of a column name, or None."""
    low = [h.strip().lower() for h in header]
    try:
        return low.index(name.strip().lower())
    except ValueError:
        return None


def _tsv_cell(row, i):
    return row[i].strip() if i is not None and i < len(row) and row[i] is not None else ""


def tsv_records(files):
    """Yield (is_primary, acct, num, desc, inst) across the given TSV files.
    files: [(path, is_primary)]."""
    for path, is_primary in files:
        header, rows = read_tsv(path)
        ai, ni = _ci_index(header, "Account"), _ci_index(header, "Account #")
        di, ii = _ci_index(header, "Description"), _ci_index(header, "Institution")
        for row in rows:
            acct, num = _tsv_cell(row, ai), _tsv_cell(row, ni)
            desc, inst = _tsv_cell(row, di), _tsv_cell(row, ii)
            if acct or num or desc or inst:
                yield is_primary, acct, num, desc, inst


def apply_tsv(in_path: Path, out_path: Path, pattern, repl, number_by_key, inst_by_key):
    """Anonymize one TSV file into out_path, keeping only the 7 important columns.
    Account # -> 'xxxx'+fake last-four (keyed on last-four); Account name +
    Description via regex; Institution -> canonical. Returns counts."""
    header, rows = read_tsv(in_path)
    idx = {name: _ci_index(header, name) for name in IMPORTANT_COLUMNS}
    out = [IMPORTANT_COLUMNS[:]]
    cells = insts = 0
    for row in rows:
        vals = {name: _tsv_cell(row, idx[name]) for name in IMPORTANT_COLUMNS}
        if not any(vals.values()):
            continue  # drop fully-empty rows (trailing junk)
        key = acct_key(vals["Account #"]) if vals["Account #"] else ""
        # Account name + Description via the regex pass.
        for tcol in ("Description", "Account"):
            if vals[tcol]:
                new = pattern.sub(repl, vals[tcol])
                if new != vals[tcol]:
                    cells += 1
                vals[tcol] = new
        # Account # normalized to xxxx+fake last-four.
        if key in number_by_key and vals["Account #"] != number_by_key[key]:
            vals["Account #"] = number_by_key[key]
            cells += 1
        # Institution -> canonical for this account.
        if key in inst_by_key and inst_by_key[key] and vals["Institution"] != inst_by_key[key]:
            vals["Institution"] = inst_by_key[key]
            insts += 1
        out.append([vals[name] for name in IMPORTANT_COLUMNS])
    write_tsv(out_path, out)
    return {"cells": cells, "institutions": insts, "rows": len(out) - 1}


def scan_rows(header, rows, word_rx, digit_rx, scan_cols=None):
    """Return leak records [(row#, col_name, cell, matched)] for a TSV table.

    Only the identifier-bearing columns are scanned: by default SCAN_COLUMNS,
    or the explicit `scan_cols` set (lower-cased column names) for bank files
    where only the designated description column carries identifiers."""
    want = scan_cols if scan_cols is not None else SCAN_COLUMNS
    cols = {i for i, h in enumerate(header) if _clean(h).lower() in want}
    leaks = []
    for r, row in enumerate(rows, start=2):
        for c in cols:
            val = row[c].strip() if c < len(row) and row[c] else ""
            if not val:
                continue
            for rx in (word_rx, digit_rx):
                if rx and rx.search(val):
                    name = header[c] if c < len(header) else f"col{c}"
                    leaks.append((r, name, val, rx.search(val).group()))
                    break
    return leaks


def count_in_rows(header, rows, matchers):
    """Count matches per (key, rx, colnames) over a TSV table."""
    wanted = set().union(*[cols for _k, _rx, cols in matchers]) if matchers else set()
    colidx = {name: _ci_index(header, name) for name in wanted}
    counts = {k: 0 for k, _rx, _c in matchers}
    for row in rows:
        for key, rx, cols in matchers:
            for cn in cols:
                i = colidx.get(cn)
                if i is not None and i < len(row) and row[i]:
                    counts[key] += len(rx.findall(row[i]))
    return counts


def run_tsv(trans_path: Path, rules_path, config_path: Path, seed, out_dir: Path,
            map_out: Path, auto_seed=False):
    """Full TSV pipeline over Transactions + (optional) rules, ONE shared mapping."""
    files = [(trans_path, True)]
    if rules_path:
        files.append((rules_path, False))

    print("=" * 64)
    print("ANONYMIZER RUN (tab-delimited)")
    print("=" * 64)
    print(f"  transactions: {trans_path}")
    print(f"  rules:        {rules_path or '(none)'}")
    print(f"  config:       {config_path}")
    print(f"  seed:         {seed}" + ("   (auto-generated - pass to seed= to reproduce)" if auto_seed else ""))

    identifiers = aggregate_identifiers(tsv_records(files))
    config = load_config(config_path)
    pools = load_pools()
    mapping = build_mapping(identifiers, config, pools, seed)
    write_mapping(mapping, map_out)
    pattern, repl, _stats = build_replacer(mapping)
    num_by_key, inst_by_key = account_output_maps(mapping)

    # Apply to every input; the SAME mapping is used for all, so the two files
    # stay perfectly in sync.
    produced = []           # (in_path, out_path)
    totals = {"cells": 0, "institutions": 0, "rows": 0}
    for path, _pri in files:
        out_path = out_dir / f"{path.stem}_anon.{seed}{path.suffix}"
        st = apply_tsv(path, out_path, pattern, repl, num_by_key, inst_by_key)
        produced.append((path, out_path))
        for k in totals:
            totals[k] += st[k]
        print(f"  -> {out_path.name}: {st['rows']} rows, {st['cells']} cells changed, "
              f"{st['institutions']} institutions normalized")
    print(f"  mapping: {map_out}  (re-identification key - keep private)")

    # ---- Checks ----
    results = []

    # Integrity: Seq/Date/Amount must pass through untouched.
    int_ok = True
    for in_path, out_path in produced:
        ih, irows = read_tsv(in_path)
        _oh, orows = read_tsv(out_path)
        irows = [r for r in irows if any(_tsv_cell(r, _ci_index(ih, n)) for n in IMPORTANT_COLUMNS)]
        for keep in ("Seq", "Date", "Amount"):
            ii = _ci_index(ih, keep)
            oi = _ci_index(IMPORTANT_COLUMNS, keep)
            before = [_tsv_cell(r, ii) for r in irows]
            after = [_tsv_cell(r, oi) for r in orows]
            if before != after:
                int_ok = False
    results.append(("Non-target integrity (Seq/Date/Amount passed through)", int_ok))

    # Leakage: nothing real may survive in the outputs (self-tested on inputs).
    word_rx, digit_rx = build_leak_scanners(identifiers, config)
    in_leaks = out_leaks = 0
    sample = []
    for in_path, out_path in produced:
        ih, irows = read_tsv(in_path)
        in_leaks += len(scan_rows(ih, irows, word_rx, digit_rx))
        oh, orows = read_tsv(out_path)
        ol = scan_rows(oh, orows, word_rx, digit_rx)
        out_leaks += len(ol)
        sample += ol[:5]
    leak_ok = out_leaks == 0 and in_leaks > 0
    results.append((f"Leakage scan (input leaks={in_leaks} -> output leaks={out_leaks})", leak_ok))
    for r, name, cell, hit in sample[:10]:
        print(f"    LEAK R{r} [{name}]: {hit!r} in {cell!r}")

    # Hit-count across BOTH outputs vs inputs.
    gated = hitcount_matchers(mapping)
    mode_by_key = {k: m for k, *_rest, m in gated}
    before = {k: 0 for k, *_ in gated}
    after = {k: 0 for k, *_ in gated}
    for in_path, out_path in produced:
        ih, irows = read_tsv(in_path)
        oh, orows = read_tsv(out_path)
        b = count_in_rows(ih, irows, [(k, rx, c) for k, rx, c, _f, _fc, _m in gated])
        a = count_in_rows(oh, orows, [(k, rx, c) for k, _r, _rc, rx, c, _m in gated])
        for k in before:
            before[k] += b[k]
            after[k] += a[k]
    bad = [(k, before[k], after[k]) for k in before
           if hitcount_failed(mode_by_key[k], before[k], after[k])]
    results.append((f"Hit-count check ({len(gated)} rules, {len(bad)} failing)", not bad))
    for k, b, a in bad[:10]:
        print(f"    FAIL {k}: before={b} after={a}")

    # Determinism: same seed -> identical mapping (apply is a pure function of it).
    map2 = build_mapping(aggregate_identifiers(tsv_records(files)), config, pools, seed)
    det_ok = json.dumps(map2, sort_keys=True) == json.dumps(mapping, sort_keys=True)
    results.append(("Determinism (same seed -> identical mapping)", det_ok))

    print("\n" + "-" * 64)
    print("REPORT")
    print("-" * 64)
    all_ok = True
    for label, passed in results:
        print(f"  [{'PASS' if passed else 'FAIL'}]  {label}")
        all_ok = all_ok and passed
    print("-" * 64)
    print("OVERALL:", "PASS - anonymized files are safe to share." if all_ok
          else "FAIL - do NOT share; see failures above.")
    print("=" * 64)
    return 0 if all_ok else 1


# --------------------------------------------------------------------------- #
# Bank CSV path — anonymize downloaded bank statements with the SAME seed's
# mapping so their fakes line up with the transactions. Reuses (and extends)
# the saved mapping; keeps each bank's own columns; touches only Description.
# --------------------------------------------------------------------------- #

def read_csv_any(path: Path):
    """Read a CSV as raw string rows, auto-detecting encoding. Returns
    (encoding, rows). Comma-delimited (bank exports)."""
    for enc in ("utf-8-sig", "cp1252"):
        try:
            with open(path, encoding=enc, newline="") as f:
                return enc, list(csv.reader(f))
        except UnicodeDecodeError:
            continue
    with open(path, encoding="cp1252", errors="replace", newline="") as f:
        return "cp1252", list(csv.reader(f))


def write_csv_any(path: Path, rows, encoding):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding=encoding, newline="") as f:
        csv.writer(f).writerows(rows)


def find_header_row(rows, desc_col):
    """Index of the transaction header row: the first row that names the
    description column AND has >=3 non-empty cells (skips summary preambles like
    BoA's, which have the word 'Description' but only a couple of columns)."""
    for i, row in enumerate(rows):
        cells = [c.strip() for c in row]
        if any(c.lower() == desc_col.lower() for c in cells) and sum(1 for c in cells if c) >= 3:
            return i
    return 0


def extend_digit_runs(mapping, descriptions, seed):
    """Add fakes for any 5+ digit runs in the bank descriptions not already in
    the mapping. New fakes are deterministic per (seed, value) so re-runs are
    stable, and distinct from existing/real values. Returns count added."""
    existing = mapping["digit_runs"]
    used = set(existing.values())
    reals = set(existing)
    new = sorted({run for d in descriptions for run in DIGIT_RUN_RE.findall(d)
                  if run not in existing})
    for real in new:
        for salt in range(1000):
            r = random.Random(f"{seed}|run|{real}|{salt}")
            cand = "".join(str(r.randrange(10)) for _ in range(len(real)))
            if cand not in used and cand != real and cand not in reals:
                existing[real] = cand
                used.add(cand)
                break
    return len(new)


def extend_alnum_codes(mapping, descriptions, seed):
    """Record fakes for letter+digit codes in the bank descriptions not already
    in the mapping (so the re-identification key stays complete). Returns count."""
    existing = mapping.setdefault("alnum_codes", {})
    new = sorted({c for d in descriptions for c in extract_codes(d) if c not in existing})
    for real in new:
        existing[real] = fake_alnum(seed, real)
    return len(new)


def leak_scanners_from_mapping(mapping, config):
    """Build (word_rx, digit_rx) leak matchers from the mapping's real side +
    config, for scanning bank output (no separate identifiers dict needed)."""
    word_terms = set(config["Blacklist"]["distinct"])
    word_terms |= set(mapping["names"]) | set(mapping["towns"])
    word_terms |= set(mapping["account_names"])
    for a in mapping["accounts"]:
        word_terms |= set(a["real_forms"])
    digit_terms = set(mapping["last_four"]) | set(mapping["digit_runs"])
    w = _alt(sorted(t for t in word_terms if t))
    word_rx = re.compile(rf"(?<!\w){w}(?!\w)", re.IGNORECASE) if w else None
    digit_rx = None
    if digit_terms:
        d = "(?:" + "|".join(re.escape(t) for t in sorted(digit_terms)) + ")"
        digit_rx = re.compile(rf"(?<![A-Za-z0-9]){d}(?![A-Za-z0-9])")
    return word_rx, digit_rx


MONEY_RE = re.compile(r"^\(?-?\$?[\d,]+\.\d{2}\)?$")

AMOUNT_COLUMN_NAMES = {"amount", "running bal.", "running balance", "balance", "summary amt."}


def parse_money(text: str):
    """Parse a plain currency string ('24,120.66', '-113,871.10', '(45.02)')
    to a float, or None if it isn't one (dates, descriptions, blanks)."""
    t = text.strip()
    if not t or not MONEY_RE.match(t):
        return None
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()").lstrip("$").replace(",", "")
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v


def format_money(value: float, like: str) -> str:
    """Format `value` as currency text, matching the thousands-comma style of
    the original text `like` it's replacing."""
    return f"{value:,.2f}" if "," in like else f"{value:.2f}"


def scale_amount_cell(text: str, factor: float) -> str:
    """Scale a currency cell by `factor`, preserving its formatting style.
    Returns `text` unchanged if it doesn't parse as plain currency."""
    v = parse_money(text)
    if v is None:
        return text
    return format_money(round(v * factor, 2), text)


def anonymize_bank_csv(in_path: Path, out_path: Path, pattern, repl, desc_col, amount_factor: float = 1.0):
    """Anonymize one bank CSV: find the real header, rewrite only the
    description column in data rows, keep every other column, drop empty padding
    rows. If amount_factor != 1.0, also scale every dollar figure - the data
    rows' Amount/Running Bal. columns by column, AND any preamble line (e.g.
    'Beginning balance as of ...') by scanning for plain currency text, since
    preamble layout isn't as predictable as the real header's. A constant
    factor preserves all the totals/running-balance math, just uniformly
    scaled. Returns a summary dict."""
    enc, rows = read_csv_any(in_path)
    hidx = find_header_row(rows, desc_col)
    header = rows[hidx]
    di = _ci_index(header, desc_col)
    amount_is = [i for i, h in enumerate(header) if h.strip().lower() in AMOUNT_COLUMN_NAMES]

    out = []
    if amount_factor != 1.0:
        for row in rows[:hidx]:                   # preamble: scan every cell
            out.append([scale_amount_cell(c, amount_factor) for c in row])
        out.append(list(header))                  # header row itself: verbatim
    else:
        out = [r for r in rows[:hidx + 1]]         # preamble + header, verbatim

    changed = kept = 0
    for row in rows[hidx + 1:]:
        if not any(c.strip() for c in row):
            continue                              # drop empty padding rows
        nr = list(row)
        if di is not None and di < len(nr) and nr[di].strip():
            new = pattern.sub(repl, nr[di])
            if new != nr[di]:
                changed += 1
            nr[di] = new
        if amount_factor != 1.0:
            for ai in amount_is:
                if ai < len(nr) and nr[ai].strip():
                    nr[ai] = scale_amount_cell(nr[ai], amount_factor)
        out.append(nr)
        kept += 1
    write_csv_any(out_path, out, enc)
    return {"out": out_path, "rows": kept, "changed": changed,
            "header_row": hidx, "desc_index": di, "encoding": enc}


def run_banks(specs, seed, config_path: Path, out_dir: Path, map_path: Path):
    """Anonymize bank CSVs using the seed's saved mapping (loaded + extended).
    specs: list of (file: Path, last4: str, desc_col: str)."""
    print("=" * 64)
    print("ANONYMIZER RUN (bank CSVs)")
    print("=" * 64)
    if not map_path.exists():
        print(f"ERROR: mapping not found: {map_path}", file=sys.stderr)
        print("Run the transactions first (that builds the mapping for this seed).", file=sys.stderr)
        return 2
    if not specs:
        print("ERROR: no bank files listed in the banks config.", file=sys.stderr)
        return 2
    missing = [str(p) for p, _l, _d in specs if not p.exists()]
    if missing:
        print("ERROR: bank file(s) not found (check the 'bank =' lines in run.config.txt):", file=sys.stderr)
        for m in missing:
            print(f"  - {m}", file=sys.stderr)
        return 2
    print(f"  seed:    {seed}")
    print(f"  mapping: {map_path}  (loaded + extended in place)")

    with open(map_path, encoding="utf-8") as f:
        mapping = json.load(f)
    config = load_config(config_path)

    # Gather all descriptions first, to extend the mapping for new codes.
    parsed = []
    all_desc = []
    for path, last4, desc_col in specs:
        enc, rows = read_csv_any(path)
        hidx = find_header_row(rows, desc_col)
        di = _ci_index(rows[hidx], desc_col)
        for row in rows[hidx + 1:]:
            if di is not None and di < len(row) and row[di].strip():
                all_desc.append(row[di])
        parsed.append((path, last4, desc_col))
    added = extend_digit_runs(mapping, all_desc, seed)
    added_alnum = extend_alnum_codes(mapping, all_desc, seed)
    if added or added_alnum:
        with open(map_path, "w", encoding="utf-8") as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False, sort_keys=True)
            f.write("\n")
    print(f"  new codes added to mapping: {added} digit-runs, {added_alnum} letter+digit")

    pattern, repl, _stats = build_replacer(mapping)
    word_rx, digit_rx = leak_scanners_from_mapping(mapping, config)

    produced = []
    total_leaks_in = total_leaks_out = 0
    for path, last4, desc_col in parsed:
        out_path = out_dir / f"{path.stem}_anon.{seed}{path.suffix}"
        info = anonymize_bank_csv(path, out_path, pattern, repl, desc_col)
        produced.append(out_path)
        # Leak self-test on input + scan output, in the description column only.
        want = {desc_col.lower()}
        _e, in_rows = read_csv_any(path)
        _e2, out_rows = read_csv_any(out_path)
        ih = in_rows[info["header_row"]] if info["header_row"] < len(in_rows) else []
        li = len(scan_rows(ih, in_rows[info["header_row"] + 1:], word_rx, digit_rx, want))
        oh = out_rows[info["header_row"]] if info["header_row"] < len(out_rows) else []
        lo = scan_rows(oh, out_rows[info["header_row"] + 1:], word_rx, digit_rx, want)
        total_leaks_in += li
        total_leaks_out += len(lo)
        tag = f" (account {last4})" if last4 else ""
        print(f"  -> {out_path.name}{tag}: {info['rows']} rows, "
              f"{info['changed']} descriptions changed, header@row{info['header_row']+1}")
        for r, name, cell, hit in lo[:5]:
            print(f"       LEAK R{r}: {hit!r} in {cell!r}")

    ok = total_leaks_out == 0 and total_leaks_in > 0
    print("\n" + "-" * 64)
    print(f"  Leakage: input leaks={total_leaks_in} -> output leaks={total_leaks_out}")
    print("OVERALL:", "PASS - bank files anonymized, consistent with the transactions."
          if ok else "FAIL - see leaks above (or self-test found nothing).")
    print("=" * 64)
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Config-file mode: run everything from a short settings file (no CLI args)
# --------------------------------------------------------------------------- #

DEFAULT_RUN_CONFIG = Path(__file__).resolve().parent / "run.config.txt"


def parse_run_config(path: Path) -> dict:
    """Parse a simple 'key = value' settings file. '#' starts a comment.

    Repeated 'bank =' lines accumulate into a list under settings['_banks']
    (each value is a '<file> | <last-four> | <column>' spec); all other keys are
    single-valued.
    """
    settings = {"_banks": []}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, val = line.split("=", 1)
        key, val = key.strip().lower(), val.strip()
        if key == "bank":
            if val:
                settings["_banks"].append(val)
        else:
            settings[key] = val
    return settings


def bank_specs_from(base: Path, lines):
    """Turn 'file | last-four | column' spec lines into (path, last4, col)."""
    specs = []
    for line in lines:
        parts = [p.strip() for p in line.split("|")]
        f = parts[0]
        last4 = parts[1] if len(parts) > 1 else ""
        col = parts[2] if len(parts) > 2 and parts[2] else "Description"
        specs.append((_resolve_path(base, f), last4, col))
    return specs


def _resolve_path(base: Path, value: str, bare_subdir: str = None) -> Path:
    """Resolve a config path: absolute as-is; a bare filename optionally into
    bare_subdir; otherwise relative to the settings file's folder."""
    p = Path(value)
    if p.is_absolute():
        return p
    if bare_subdir and p.parent == Path("."):
        return base / bare_subdir / p.name
    return base / p


def cmd_fromconfig(args) -> int:
    cfg_file: Path = args.file or DEFAULT_RUN_CONFIG
    if not cfg_file.exists():
        print(f"ERROR: settings file not found: {cfg_file}", file=sys.stderr)
        print("Create it (see run.config.txt) with: input, seed, output, config.", file=sys.stderr)
        return 2

    base = cfg_file.parent
    s = parse_run_config(cfg_file)
    banks_only = getattr(args, "banks_only", False)
    print(f"Reading settings from: {cfg_file}"
          + ("   (banks only)" if banks_only else "") + "\n")

    if not s.get("config"):
        print("ERROR: 'config' (Names/Towns/Blacklist file) is required in the settings file.", file=sys.stderr)
        return 2
    config_path = _resolve_path(base, s["config"])
    out_dir = _resolve_path(base, s["output_dir"]) if s.get("output_dir") else (base / "output")
    seed = s.get("seed") or secrets.token_hex(3)
    auto = not s.get("seed")
    map_out = MAPPING_DIR / f"mapping.{seed}.json"
    bank_specs = bank_specs_from(base, s.get("_banks", []))

    # Banks-only: skip transactions, reuse the existing seed's mapping.
    if banks_only:
        if not bank_specs:
            print("ERROR: no 'bank =' lines in the settings file.", file=sys.stderr)
            return 2
        return run_banks(bank_specs, seed, config_path, out_dir, map_out)

    # Full TSV run: transactions (+ rules), then any bank CSVs, one shared mapping.
    if s.get("transactions"):
        trans = _resolve_path(base, s["transactions"])
        rules = _resolve_path(base, s["rules"]) if s.get("rules") else None
        for label, p in [("transactions", trans), ("rules", rules), ("config", config_path)]:
            if p and not p.exists():
                print(f"ERROR: {label} file not found: {p}", file=sys.stderr)
                return 2
        rc = run_tsv(trans, rules, config_path, seed, out_dir, map_out, auto_seed=auto)
        if bank_specs:
            print()
            rc = run_banks(bank_specs, seed, config_path, out_dir, map_out) or rc
        return rc

    # No transactions but bank files listed -> banks only (reuse mapping).
    if bank_specs:
        return run_banks(bank_specs, seed, config_path, out_dir, map_out)

    # Legacy xlsx mode: an 'input' workbook is listed.
    if not s.get("input"):
        print("ERROR: settings file needs 'transactions' (+ optional 'bank =' lines) or 'input' (xlsx).",
              file=sys.stderr)
        return 2
    sheets = [x for x in re.split(r"[,\s]+", s["sheets"]) if x] if s.get("sheets") else None
    ns = argparse.Namespace(
        input=_resolve_path(base, s["input"]),
        config=config_path,
        seed=(s.get("seed") or None),
        out=(_resolve_path(base, s["output"], bare_subdir="output") if s.get("output") else None),
        map_out=None,
        sheets=sheets,
    )
    return cmd_run(ns)


# --------------------------------------------------------------------------- #
# Draft 9: folder-level Excel + CSV anonymizer (extends the TSV core above).
# Reads the master workbook READ-ONLY (never rewrites it), one shared seed for
# the whole folder. Increment 1 = load + no-op round-trip (this gates the rest).
# --------------------------------------------------------------------------- #

def used_width(row_values) -> int:
    """Column count up to the last non-empty cell — clamps phantom columns
    (RulesN reports 16k columns but really has ~17)."""
    last = 0
    for i, v in enumerate(row_values, start=1):
        if v is not None and str(v).strip() != "":
            last = i
    return last


def inspect_workbook9(path: Path):
    """Open the workbook read-only and describe every sheet: state, reported vs
    real (clamped) column count, header, and whether it's a data tab to process
    (visible sheets) or skipped (hidden/veryHidden)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tabs = []
    for ws in wb.worksheets:
        row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()) or ())
        width = used_width(row1)
        tabs.append({
            "title": ws.title,
            "state": ws.sheet_state,               # visible / hidden / veryHidden
            "reported_cols": ws.max_column or 0,
            "used_cols": width,
            "reported_rows": ws.max_row or 0,
            "header": [_clean(v) for v in row1[:width]],
            "processed": ws.sheet_state == "visible",
        })
    wb.close()
    return tabs


# A compound account string:  "<name> - <xxxxNNNN> (<4-char code>)"
COMPOUND_RE = re.compile(r"^(.*?) - (\S+) \(([^)]+)\)\s*$")


def _tab_rows(wb, title):
    """Yield (col_index_by_lowername, row_tuple) for each data row of a tab,
    using the clamped header."""
    ws = wb[title]
    it = ws.iter_rows(values_only=True)
    row1 = list(next(it, ()) or [])
    width = used_width(row1)
    idx = {_clean(v).lower(): i for i, v in enumerate(row1[:width]) if _clean(v)}
    for r in it:
        yield idx, r


def _val(idx, r, name):
    i = idx.get(name.lower())
    if i is None or i >= len(r) or r[i] is None:
        return ""
    return str(r[i]).strip()


def _find_key_tab(wb):
    """Find the relationship-key tab (Address | Prop | LLC | bank1-3) by its
    COLUMNS, not a fixed name. The founder has already renamed it once
    ('LLCs' -> 'Properties'); hardcoding the name means a future rename
    silently harvests nothing - no error, just corrupted output (addresses
    left unparsed, ZIPs caught by the generic digit scrub, property codes
    decoupled from their real address) with nothing in the report to say why."""
    needed = {"address", "prop", "llc"}
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        row1 = next(ws.iter_rows(values_only=True), ())
        header = {_clean(v).lower() for v in row1}
        if needed <= header:
            return ws.title
    return None


def harvest9(path: Path) -> dict:
    """Harvest structural identifiers (no config) from the data tabs.
    Returns sets: account_names, account_numbers, entity_codes, llc_names,
    property_codes, addresses, compounds; plus prop_addresses/llc_addresses
    (dicts: property code / LLC name -> set of addresses it's paired with in
    the key tab), used to derive the property-code and LLC-name fakes from
    that property's own fake street."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    acct_names, acct_nums, codes = set(), set(), set()
    llcs, props, addrs, compounds = set(), set(), set(), set()
    prop_addrs = {}
    llc_addrs = {}

    def add_compound(s):
        if not s:
            return
        compounds.add(s)
        m = COMPOUND_RE.match(s)
        if m:
            nm, num, code = (g.strip() for g in m.groups())
            if nm:
                acct_names.add(nm)
            if num:
                acct_nums.add(num)
            if code:
                codes.add(code)

    present = set(wb.sheetnames)
    if "accounts" in present:
        for idx, r in _tab_rows(wb, "accounts"):
            acct_names.add(_val(idx, r, "Account")) if _val(idx, r, "Account") else None
            acct_nums.add(_val(idx, r, "Account #")) if _val(idx, r, "Account #") else None
            llcs.add(_val(idx, r, "LLC")) if _val(idx, r, "LLC") else None
            add_compound(_val(idx, r, "Unique Account Identifier"))
            add_compound(_val(idx, r, "Sorted Accounts"))
    key_tab = _find_key_tab(wb)
    if key_tab:
        for idx, r in _tab_rows(wb, key_tab):
            addr = _val(idx, r, "Address")
            prop = _val(idx, r, "Prop")
            if addr:
                addrs.add(addr)
            if prop:
                props.add(prop)
            if prop and addr:
                prop_addrs.setdefault(prop.lower(), set()).add(addr)
            llc = _val(idx, r, "LLC")
            if llc:
                llcs.add(llc)
                if addr:
                    llc_addrs.setdefault(llc.lower(), set()).add(addr)
            for b in ("bank1", "bank2", "bank3"):
                if _val(idx, r, b):
                    acct_nums.add(_val(idx, r, b))
    # Account / Account # / Prop also appear on the transaction-style tabs.
    for tab in ("Transactions", "RulesN"):
        if tab in present:
            for idx, r in _tab_rows(wb, tab):
                acct_names.add(_val(idx, r, "Account")) if _val(idx, r, "Account") else None
                acct_nums.add(_val(idx, r, "Account #")) if _val(idx, r, "Account #") else None
                props.add(_val(idx, r, "Prop")) if _val(idx, r, "Prop") else None
    wb.close()
    return {
        "account_names": acct_names - {""}, "account_numbers": acct_nums - {""},
        "entity_codes": codes - {""}, "llc_names": llcs - {""},
        "property_codes": props - {""}, "addresses": addrs - {""},
        "compounds": compounds - {""}, "prop_addresses": prop_addrs,
        "llc_addresses": llc_addrs, "key_tab": key_tab,
    }


def cmd_harvest9(args) -> int:
    src: Path = args.workbook
    if not src.exists():
        print(f"ERROR: workbook not found: {src}", file=sys.stderr)
        return 2
    print("Draft 9 - increment 2: harvest structural identifiers + config inventory")
    print(f"  workbook: {src}")
    h = harvest9(src)

    def show(label, values, n=12):
        vals = sorted(values, key=str.lower)
        print(f"\n{label}: {len(vals)} distinct")
        for v in vals[:n]:
            print(f"    {v!r}")
        if len(vals) > n:
            print(f"    ... (+{len(vals) - n} more)")

    show("Account names (harvested)", h["account_names"])
    show("Account numbers (harvested)", h["account_numbers"])
    show("Entity/4-char codes (harvested)", h["entity_codes"])
    show("LLC/entity names (harvested)", h["llc_names"])
    show("Property codes (harvested)", h["property_codes"])
    show("Addresses (harvested)", h["addresses"])
    show("Compound account strings (harvested)", h["compounds"])

    if args.config and args.config.exists():
        cfg = load_config(args.config)
        print("\nConfig CSV:")
        for name in CONFIG_COLUMNS:
            print(f"  {name}: {len(cfg[name]['distinct'])} distinct")

    print("\nRESULT: PASS - distinct real-value inventory harvested (detection only; no mapping yet).")
    return 0


MAPPINGS9_DIR = Path(__file__).resolve().parent / "mappings"   # per-seed: mappings/<seed>/


def parse_address(addr: str):
    """Split 'street, city, ST ZIP' into (street, city, state, zip). Tolerant of
    the two comma styles seen ('..., NJ, 07026' and '..., NJ 07047')."""
    parts = [p.strip() for p in addr.split(",") if p.strip()]
    street = parts[0] if parts else ""
    city = parts[1] if len(parts) > 1 else ""
    z = re.search(r"\b(\d{5})\b", addr)
    st = re.search(r"\b([A-Za-z]{2})\b(?=[,\s]*\d{5})", addr) or re.search(r",\s*([A-Za-z]{2})\b", addr)
    return street, city, (st.group(1) if st else ""), (z.group(1) if z else "")


def build_mapping9(harvest: dict, config: dict, pools: dict, seed) -> dict:
    """One deterministic, seed-driven, BIJECTIVE mapping across every category.
    A single global `used` set guarantees distinct reals -> distinct fakes so the
    LLC key joins survive. Names/LLCs are case-insensitive (one fake per value)."""
    rng = random.Random(seed)
    used = set()

    cfg_names = set(config["Names"]["distinct"])
    cfg_towns = set(config["Towns"]["distinct"])
    cfg_black = set(config["Blacklist"]["distinct"])

    # Forbid fakes that equal or CONTAIN any real root (so the blacklist scan
    # can't be tripped by a pool word like the surname "Michael").
    forbidden_words = (cfg_black | cfg_names | cfg_towns
                       | harvest["account_names"] | harvest["llc_names"])
    fw = _alt(sorted(w for w in forbidden_words if w))
    forbid_rx = re.compile(rf"{WORDISH_PRE}{fw}{WORDISH_POST}", re.IGNORECASE) if fw else None

    def contains_forbidden(t):
        return bool(forbid_rx and forbid_rx.search(t))

    all_reals = (cfg_names | cfg_towns | harvest["account_names"] | harvest["llc_names"]
                 | harvest["property_codes"] | harvest["entity_codes"] | harvest["addresses"])

    def draw(make, reject=None):
        for _ in range(10000):
            c = make()
            if c and c not in used and c not in all_reals and not (reject and reject(c)):
                used.add(c)
                return c
        raise RuntimeError("could not draw a unique fake; pool too small?")

    def fake_person():
        return f"{rng.choice(pools['first_names'])} {rng.choice(pools['surnames'])}"

    def fake_single():
        return rng.choice(pools["first_names"])

    def unique_code(real):
        for salt in range(10000):
            cand = fake_alnum(f"{seed}:{salt}" if salt else seed, real)
            if cand != real and cand not in used:
                used.add(cand)
                return cand
        raise RuntimeError("code space exhausted")

    # --- Names: ONLY the config's Names column, case-insensitive ---
    # Account/LLC names harvested structurally (e.g. 'Realty Central CK',
    # 'MR Business 3', 'Blue Cash Everyday') are NOT auto-faked just for
    # appearing in the Account/LLC column - only the specific words the
    # founder has actually listed in Names get replaced, everywhere they
    # appear (Account column, LLC column, Transactions/notes text alike).
    # A real entity root already in Names (e.g. 'Danmir') still correctly
    # covers ALL its compound forms ('Danmir CK', 'Danmir LLC', 'Danmir Bus')
    # via plain whole-word matching - the suffix is untouched literal text,
    # so one config entry per root is enough; no separate harvesting needed.
    names = {}
    name_reals = {}
    for r in sorted(cfg_names):
        name_reals.setdefault(r.lower(), r)   # sorted -> deterministic representative

    def split_core(real):
        parts = real.split()
        if len(parts) > 1 and parts[-1].upper().strip(".,") in ACCOUNT_TYPE_WORDS:
            return parts[:-1], parts[-1]      # (core words, suffix word)
        return parts, ""

    core_fakes = {}                           # core (lowercased) -> shared fake base
    def base_for(core):
        key = " ".join(core).lower()
        if key not in core_fakes:
            maker = fake_single if len(core) <= 1 else fake_person
            core_fakes[key] = draw(maker, reject=contains_forbidden)
        return core_fakes[key]

    for low in sorted(name_reals):
        core, suffix = split_core(name_reals[low])
        base = base_for(core)
        names[low] = f"{base} {suffix}" if suffix else base

    # --- Towns (config) ---
    towns = {}
    for r in sorted(cfg_towns):
        towns[r.lower()] = draw(lambda: rng.choice(pools["towns"]), reject=contains_forbidden)

    # --- Account numbers, keyed on last-four ---
    accounts, last_four = [], {}
    by_key = {}
    for num in harvest["account_numbers"]:
        by_key.setdefault(acct_key(num), set()).add(num)
    for key in sorted(by_key):
        flf = draw(lambda: f"{rng.randrange(10000):04d}")
        for form in bare_forms(key):
            last_four[form] = flf
        accounts.append({"key": key, "real_forms": sorted(by_key[key]),
                         "fake_last4": flf, "fake_number": "xxxx" + flf})

    # --- Entity 4-char codes (shape-preserving, unique) ---
    entity_codes = {c: unique_code(c) for c in sorted(harvest["entity_codes"])}

    # --- Streets (from addresses) + addresses ---
    # Street numbers are capped to 1-999 (1-3 digits): property codes are
    # derived FROM these fake street numbers (see below) and must never need
    # truncation - "up to 3 digits" is naturally satisfied by this range.
    streets, addresses = {}, {}
    street_pool = pools.get("streets", ["Main St"])

    def fake_street():
        return f"{rng.randrange(1, 1000)} {rng.choice(street_pool)}"
    for addr in sorted(harvest["addresses"]):
        street, city, state, zipc = parse_address(addr)
        if street and street.lower() not in streets:
            streets[street.lower()] = draw(fake_street)
        fstreet = streets.get(street.lower(), street)
        fcity = towns.get(city.lower())
        if city and fcity is None:                    # city not in config towns
            fcity = draw(lambda: rng.choice(pools["towns"]))
            towns[city.lower()] = fcity
        addresses[addr] = ", ".join(x for x in
                                    [fstreet, fcity or city, f"{state} {zipc}".strip()] if x)

    # --- Property codes: derived from the property's own fake street ---
    # Fake = up to 3 digits from the fake street number + '-' + 3 letters
    # from the fake street name (e.g. fake '142 Maple Ave' -> '142-MAP';
    # fake '58 Oak Ln' -> '58-OAK'; a number under 3 digits is used as-is).
    # Only values that actually LOOK like a property code (contain a digit,
    # e.g. 'G42VW') are converted at all - the 'Prop' column also carries
    # plain administrative labels ('Commo', 'UPDATE', 'FILL') that are left
    # untouched. A stray real town/name with no digits is NOT caught here -
    # add it to Names/Towns/Blacklist in the config if it needs protecting.
    def prop_code_from_street(fake_street_text):
        parts = fake_street_text.split(maxsplit=1)
        number = parts[0] if parts else ""
        name_word = parts[1].split()[0] if len(parts) > 1 and parts[1].split() else ""
        digits = number[:3]
        letters = "".join(c for c in name_word if c.isalpha())[:3].upper()
        return f"{digits}-{letters}"

    def synth_property_street(real_code):
        # A property code with no linked Address in LLCs still needs a fake
        # street to derive its code from; seeded on the real code itself so
        # it's stable regardless of harvest order (never shown as a real
        # address anywhere - it only exists to derive this one code).
        r = random.Random(f"{seed}|propstreet|{real_code}")
        return f"{r.randrange(1, 1000)} {r.choice(street_pool)}"

    addr_for_prop = {p: sorted(a)[0] for p, a in harvest["prop_addresses"].items()}
    property_codes = {}
    for real in sorted(harvest["property_codes"]):     # full sort -> deterministic
        low = real.lower()
        if low in property_codes or not any(c.isdigit() for c in real):
            continue
        fstreet = None
        addr = addr_for_prop.get(low)
        if addr:
            street, _city, _state, _zipc = parse_address(addr)
            fstreet = streets.get(street.lower())
        if not fstreet:
            fstreet = synth_property_street(real)
        candidate = prop_code_from_street(fstreet)
        if candidate in used:
            # Collision: two properties' fake streets produced the same
            # NNN-LLL (e.g. '142 Maple' vs '142 Maplewood' both -> '142-MAP').
            # Redraw a fresh, still street-shaped code deterministically tied
            # to THIS real code, so the mapping stays reproducible.
            for salt in range(1, 10000):
                rr = random.Random(f"{seed}|propcode|{real}|{salt}")
                candidate = (f"{rr.randrange(1, 1000)}-"
                             f"{''.join(chr(65 + rr.randrange(26)) for _ in range(3))}")
                if candidate not in used:
                    break
            else:
                raise RuntimeError("property code space exhausted")
        used.add(candidate)
        property_codes[low] = candidate

    # --- LLC names for property-owning entities: match the property's street.
    # e.g. real 'Danmir LLC' (linked to '196 Central Ave') -> fake
    # '280 Water St LLC', using that SAME property's already-generated fake
    # street. Overrides the generic owner-root fake for this EXACT compound
    # string only - the same owner's account nicknames (e.g. 'Danmir CK')
    # still use the owner-root fake, since those are a different real-world
    # reference, not this property-holding entity's registered name. An LLC
    # tied to more than one address (data can have this) is named after the
    # alphabetically-first one, for a reproducible, deterministic choice.
    def synth_llc_street(real_name):
        r = random.Random(f"{seed}|llcstreet|{real_name}")
        return f"{r.randrange(1, 1000)} {r.choice(street_pool)}"

    addr_for_llc = {n: sorted(a)[0] for n, a in harvest["llc_addresses"].items()}
    llc_property_names = {}
    for low in sorted(harvest["llc_addresses"]):       # full sort -> deterministic
        fstreet = None
        addr = addr_for_llc.get(low)
        if addr:
            street, _city, _state, _zipc = parse_address(addr)
            fstreet = streets.get(street.lower())
        if not fstreet:
            fstreet = synth_llc_street(low)
        candidate = f"{fstreet} LLC"
        if candidate in used:
            # Collision: two LLCs' fake streets produced the same name.
            # Redraw, deterministically tied to THIS real LLC name.
            for salt in range(1, 10000):
                rr = random.Random(f"{seed}|llcname|{low}|{salt}")
                candidate = f"{rr.randrange(1, 1000)} {rr.choice(street_pool)} LLC"
                if candidate not in used:
                    break
            else:
                raise RuntimeError("LLC-name space exhausted")
        used.add(candidate)
        llc_property_names[low] = candidate

    # An owner root that owns a specific property (e.g. 'Motia' -> 'Motia
    # LLC' -> '301 Brook St LLC') gets its NAME itself overridden to match,
    # not just the exact 'Motia LLC' compound - otherwise 'Motia CK' would
    # still show the old owner-name fake ('Mark CK') while 'Motia LLC' shows
    # the property-derived one, inconsistent. Every suffixed form (bare, CK,
    # Bus, ...) shares this one root, so overriding it here covers all of
    # them at once, e.g. 'Motia CK' -> '301 Brook St LLC CK'. This WINS over
    # whatever the config-driven name section assigned that root.
    for full_llc, fake_llc in llc_property_names.items():
        root, _suffix = split_core(full_llc)
        root_key = " ".join(root).lower()
        if root_key:
            names[root_key] = fake_llc

    return {
        "seed": seed, "names": names, "towns": towns, "accounts": accounts,
        "last_four": last_four, "entity_codes": entity_codes,
        "property_codes": property_codes, "streets": streets, "addresses": addresses,
        "llc_property_names": llc_property_names,
    }


def cmd_map9(args) -> int:
    src: Path = args.workbook
    if not src.exists():
        print(f"ERROR: workbook not found: {src}", file=sys.stderr)
        return 2
    if not args.config or not args.config.exists():
        print("ERROR: --config (Names/Towns/Blacklist CSV) is required.", file=sys.stderr)
        return 2
    print("Draft 9 - increment 3: build ONE deterministic bijective mapping")
    print(f"  workbook: {src}\n  seed:     {args.seed}")

    harvest = harvest9(src)
    config = load_config(args.config)
    pools = load_pools()
    mapping = build_mapping9(harvest, config, pools, args.seed)

    out = MAPPINGS9_DIR / str(args.seed) / "mapping.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False, sort_keys=True)
        f.write("\n")

    # Counts + bijection check (all fakes distinct across every category).
    cats = {"names": list(mapping["names"].values()),
            "towns": list(mapping["towns"].values()),
            "account numbers": [a["fake_number"] for a in mapping["accounts"]],
            "entity codes": list(mapping["entity_codes"].values()),
            "property codes": list(mapping["property_codes"].values()),
            "streets": list(mapping["streets"].values()),
            "addresses": list(mapping["addresses"].values())}
    all_fakes = [f for v in cats.values() for f in v]
    print(f"\n  mapping written: {out}")
    for k, v in cats.items():
        print(f"    {k:16s}: {len(v)}")
    bij = len(all_fakes) == len(set(all_fakes))
    print(f"\n  Bijection (all {len(all_fakes)} fakes distinct): {'OK' if bij else 'FAIL - collision!'}")

    print("\n  Samples:")
    for low, fake in list(mapping["names"].items())[:3]:
        print(f"    name   {low!r} -> {fake!r}")
    for a in mapping["accounts"][:2]:
        print(f"    acct#  {a['real_forms']} -> {a['fake_number']!r}")
    for r, f in list(mapping["property_codes"].items())[:2]:
        print(f"    prop   {r!r} -> {f!r}")
    for r, f in mapping["addresses"].items():
        print(f"    addr   {r!r} -> {f!r}")

    print("\nRESULT:", "PASS - deterministic bijective mapping built and written."
          if bij else "FAIL - bijection collision (see above).")
    return 0 if bij else 1


def run9_batch(src: Path, config_path: Path, seed: str, bank_specs, out: Path = None,
               auto_seed: bool = False, amount_factor: float = 1.0) -> int:
    """Draft 9 increment 7 core: one seed -> workbook tabs + bank CSVs + checks.
    bank_specs: list of (Path, desc_col). Shared by the CLI (run9) and the
    config-file front-end (fromconfig9). amount_factor (0.80-1.20, default
    1.0 = off) uniformly scales every dollar amount across the workbook and
    bank CSVs - an OPT-IN extra layer of obfuscation on top of value mapping;
    a constant factor preserves all totals/running-balance math."""
    if not (0.80 <= amount_factor <= 1.20):
        print(f"ERROR: amount_factor must be between 0.80 and 1.20 (got {amount_factor}).",
              file=sys.stderr)
        return 2
    problems = []
    if not src.exists():
        problems.append(("workbook", src))
    if not (config_path and config_path.exists()):
        problems.append(("config", config_path))
    problems += [("bank", p) for p, _c in bank_specs if not p.exists()]
    if problems:
        print("ERROR: cannot find the following file(s) - check the paths in the settings file:",
              file=sys.stderr)
        for label, p in problems:
            print(f"  [{label}] {p}", file=sys.stderr)
            parent = p.parent
            if parent.exists():
                near = sorted(x.name for x in parent.iterdir() if x.is_file())[:8]
                print(f"          folder exists; it contains: {near}"
                      + (" ..." if len(near) == 8 else ""), file=sys.stderr)
            else:
                print(f"          (that folder does not exist: {parent})", file=sys.stderr)
        return 2

    print("=" * 64)
    print(f"ANONYMIZER RUN v{VERSION} (Draft 9 - workbook + bank CSVs)")
    print("=" * 64)
    print(f"  workbook: {src}")
    print(f"  banks:    {[p.name for p, _c in bank_specs] or '(none)'}")
    print(f"  seed:     {seed}" + ("   (auto-generated)" if auto_seed else ""))
    if amount_factor != 1.0:
        print(f"  amount factor: {amount_factor}  (every dollar amount scaled by this factor)")

    harvest = harvest9(src)
    if not harvest.get("key_tab"):
        print("  WARNING: no relationship-key tab found (need a visible tab with"
              " Address + Prop + LLC columns). Addresses/property-code links will"
              " be MISSING and any address text will NOT be properly anonymized -"
              " check the workbook for a renamed or missing key tab.")
    config = load_config(config_path)
    pools = load_pools()
    mapping = build_mapping9(harvest, config, pools, seed)
    map_out = MAPPINGS9_DIR / str(seed) / "mapping.json"
    map_out.parent.mkdir(parents=True, exist_ok=True)
    with open(map_out, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False, sort_keys=True)
        f.write("\n")

    out_dir = (out or (src.parent / "output")) / str(seed)
    # Clear any files left from an earlier run under this seed (e.g. from a
    # workbook/bank file that has since been renamed or removed) - otherwise
    # the blacklist scan below would sweep them in and flag leaks that belong
    # to old, no-longer-relevant output rather than this run's files.
    if out_dir.exists():
        shutil.rmtree(out_dir)

    # Workbook tabs -> one plain .xlsx (four tabs).
    written, wb_out, stats = apply_workbook9(src, mapping, out_dir, amount_factor=amount_factor)
    print(f"\n  output: {out_dir}")
    print(f"    {wb_out.name}  ({len(written)} tabs)")
    for title, n in written:
        print(f"      - {title}: {n} rows")

    # Bank CSVs (same mapping/replacer; only the description column changes).
    pattern, repl, _bs = build_replacer9(mapping)
    for b, desc_col in bank_specs:
        safe_stem = anonymize_filename_stem(b.stem, mapping)
        out_name = f"{safe_stem}_anon{b.suffix}"
        info = anonymize_bank_csv(b, out_dir / out_name, pattern, repl, desc_col,
                                   amount_factor=amount_factor)
        print(f"    {out_name}: {info['rows']} rows, "
              f"{info['changed']} descriptions changed, header@row{info['header_row'] + 1}")
    print(f"  mapping: {map_out}  (re-identification key - keep private)")

    # Checks.
    results = []
    leaks = scan_blacklist9(out_dir, config["Blacklist"]["distinct"])
    results.append((f"Blacklist scan (whole-word) - {len(leaks)} leak(s)", not leaks))
    for fn, t, ctx in leaks[:10]:
        print(f"    LEAK [{fn}] {t!r} in ...{ctx}...")

    m2 = build_mapping9(harvest9(src), load_config(config_path), pools, seed)
    det = json.dumps(m2, sort_keys=True) == json.dumps(mapping, sort_keys=True)
    results.append(("Determinism (same seed -> identical mapping)", det))

    all_fakes = ([a["fake_number"] for a in mapping["accounts"]]
                 + list(mapping["names"].values()) + list(mapping["property_codes"].values())
                 + list(mapping["entity_codes"].values()) + list(mapping["streets"].values()))
    results.append(("Bijection (distinct reals -> distinct fakes)", len(all_fakes) == len(set(all_fakes))))

    print("\n" + "-" * 64)
    print("REPORT")
    print("-" * 64)
    ok = True
    for label, passed in results:
        print(f"  [{'PASS' if passed else 'FAIL'}]  {label}")
        ok = ok and passed
    print("-" * 64)
    print("OVERALL:", "PASS - batch anonymized, safe to share." if ok
          else "FAIL - do NOT share; see failures above.")
    print("=" * 64)
    return 0 if ok else 1


def cmd_run9(args) -> int:
    """Draft 9 increment 7 (CLI): master workbook + --banks CSVs under one seed."""
    bank_specs = [(Path(b), "Description") for b in (args.banks or [])]
    return run9_batch(args.workbook, args.config, args.seed, bank_specs, args.out,
                       amount_factor=args.amount_factor)


def cmd_fromconfig9(args) -> int:
    """Draft 9 increment 7 (double-click front-end): read every setting from a
    'key = value' file so the user never types the long command line."""
    cfg_file: Path = args.file or (Path(__file__).resolve().parent / "run9.config.txt")
    if not cfg_file.exists():
        print(f"ERROR: settings file not found: {cfg_file}", file=sys.stderr)
        print("Create it (see run9.config.example.txt): workbook, config, seed, bank = ...",
              file=sys.stderr)
        return 2
    base = cfg_file.parent
    s = parse_run_config(cfg_file)
    print(f"Reading settings from: {cfg_file}\n")

    if not s.get("config"):
        print("ERROR: 'config' (Names/Towns/Blacklist CSV) is required in the settings file.", file=sys.stderr)
        return 2
    config_path = _resolve_path(base, s["config"])
    seed = s.get("seed") or secrets.token_hex(3)
    auto = not s.get("seed")
    out = _resolve_path(base, s["output_dir"]) if s.get("output_dir") else None

    amount_factor = 1.0
    if s.get("amount_factor"):
        try:
            amount_factor = float(s["amount_factor"])
        except ValueError:
            print(f"ERROR: 'amount_factor' must be a number (got {s['amount_factor']!r}).",
                  file=sys.stderr)
            return 2

    if s.get("input_dir"):
        # --- Folder mode: point at ONE folder; take everything from it. ---
        src, bank_specs, err = discover_inputs9(_resolve_path(base, s["input_dir"]), config_path)
        if err:
            print(f"ERROR: {err}", file=sys.stderr)
            return 2
        print(f"  input folder: {_resolve_path(base, s['input_dir'])}")
        print(f"    workbook:  {src.name}")
        print(f"    banks:     {[p.name for p, _c in bank_specs] or '(none found)'}\n")
    else:
        # --- Explicit-file mode: 'workbook =' + 'bank =' lines. ---
        if not s.get("workbook"):
            print("ERROR: give 'input_dir' (a folder) OR 'workbook' (a master .xlsx/.xlsm).",
                  file=sys.stderr)
            return 2
        src = _resolve_path(base, s["workbook"])
        # 'bank = <file> | <last-four> | <column>' lines; last-four is ignored
        # here (statements are single-account), column defaults to Description.
        bank_specs = []
        for line in s.get("_banks", []):
            parts = [p.strip() for p in line.split("|")]
            f = _resolve_path(base, parts[0])
            col = parts[2] if len(parts) > 2 and parts[2] else "Description"
            bank_specs.append((f, col))

    return run9_batch(src, config_path, seed, bank_specs, out, auto_seed=auto,
                       amount_factor=amount_factor)


def discover_inputs9(folder: Path, config_path: Path):
    """Folder mode: find the one master workbook (.xlsx/.xlsm) and every bank CSV
    in `folder` (top level only). Excludes the config CSV, Excel lock files
    (~$...), and previously-produced *_anon.* outputs.
    Returns (workbook, bank_specs, error_message-or-None)."""
    if not folder.is_dir():
        return None, [], f"input folder not found: {folder}"

    def usable(p):
        return (p.is_file() and not p.name.startswith("~$")
                and not p.stem.endswith("_anon"))
    cfg_resolved = config_path.resolve()

    wbs = sorted(p for p in folder.iterdir()
                 if usable(p) and p.suffix.lower() in (".xlsx", ".xlsm"))
    if not wbs:
        return None, [], f"no .xlsx/.xlsm workbook found in {folder}"
    if len(wbs) > 1:
        names = ", ".join(p.name for p in wbs)
        return None, [], (f"{len(wbs)} workbooks found in {folder} ({names}); "
                          f"keep one, or add a 'workbook =' line to pick it.")

    banks = sorted(p for p in folder.iterdir()
                   if usable(p) and p.suffix.lower() == ".csv"
                   and p.resolve() != cfg_resolved)
    return wbs[0], [(p, "Description") for p in banks], None


def build_replacer9(mapping: dict):
    """Combined case-insensitive replacer for the Draft 9 mapping. Group order is
    most-specific first (street phrases and account numbers before names, codes
    before the generic scrubs). Returns (pattern, repl, stats)."""
    seed = mapping.get("seed", "")
    street_map = {r.lower(): f for r, f in mapping["streets"].items()}
    name_map = dict(mapping["names"])                       # already lower-keyed
    town_map = dict(mapping["towns"])
    prop_map = dict(mapping["property_codes"])              # lower-keyed
    ent_map = {k.lower(): v for k, v in mapping["entity_codes"].items()}
    llcprop_map = dict(mapping.get("llc_property_names", {}))  # already lower-keyed
    last4_map = dict(mapping["last_four"])
    acctnum_map = {}
    for a in mapping["accounts"]:
        for form in a["real_forms"]:
            # Only MIXED letter+digit forms ('xxxx7352') go in the global
            # blanket map - that shape essentially never collides with
            # ordinary text. Pure-digit forms ('7352') are handled by the
            # separate 'last4' bare-token group below. A purely alphabetic
            # account identifier (e.g. 'cash', 'PP' - no real bank number,
            # so acct_key() fell back to the raw text) is NOT safe to
            # blanket-match anywhere in the workbook: 'cash' is also an
            # ordinary word (e.g. the card name 'Blue Cash Everyday'). Those
            # are excluded here and handled by apply_workbook9 instead,
            # scoped to that account's own Account/Account # cells only.
            if any(c.isalpha() for c in form) and any(c.isdigit() for c in form):
                acctnum_map[form.lower()] = a["fake_number"]

    # Two names glued with NO separator at all (e.g. 'ofirtali' = 'ofir' +
    # 'tali') can't be caught by any whole-word boundary rule - there is no
    # character marking where one name ends and the next begins. Detect it
    # instead by checking whether the WHOLE word is exactly the
    # concatenation of two known single-word config Names, and replace it
    # with their two fakes glued the same way. Each half must be >=3 chars
    # (skips noise from very short entries like 'MR'); longest-prefix-first
    # avoids preferring a spurious short split when a longer one also fits.
    single_names = {k: v for k, v in name_map.items() if " " not in k}
    single_names_by_len = sorted(single_names, key=len, reverse=True)

    def find_glued_pair(word_low):
        if word_low in single_names:
            return None                   # already one whole known name
        for first in single_names_by_len:
            if len(first) < 3 or not word_low.startswith(first):
                continue
            second = word_low[len(first):]
            if len(second) >= 3 and second in single_names:
                return first, second
        return None

    parts = []
    def grp(gname, literals, pre=WORDISH_PRE, post=WORDISH_POST):
        body = _alt(literals)
        if body:
            parts.append(rf"(?P<{gname}>{pre}{body}{post})")
    grp("street", list(street_map))
    grp("acctnum", list(acctnum_map))
    grp("llcprop", list(llcprop_map))         # before 'name': whole compound wins
    grp("name", list(name_map))
    grp("town", list(town_map))
    grp("propcode", list(prop_map), pre=r"(?<![A-Za-z0-9])", post=r"(?![A-Za-z0-9])")
    grp("entcode", list(ent_map), pre=r"(?<![A-Za-z0-9])", post=r"(?![A-Za-z0-9])")
    if last4_map:
        l4 = "(?:" + "|".join(re.escape(k) for k in sorted(last4_map)) + ")"
        parts.append(rf"(?P<last4>(?<![A-Za-z0-9]){l4}(?![A-Za-z0-9]))")
    if single_names:
        # Letters-only run, tried before 'alcode' (which would otherwise
        # structurally claim the same span, always as a no-op for pure
        # letters, and never let 'gluedpair' see it).
        parts.append(rf"(?P<gluedpair>{WORDISH_PRE}[A-Za-z]{{4,}}{WORDISH_POST})")
    parts.append(r"(?P<run>(?<![A-Za-z0-9])\d{5,}(?![A-Za-z0-9]))")
    parts.append(r"(?P<alcode>(?<![A-Za-z0-9])[A-Za-z0-9]{5,}(?![A-Za-z0-9]))")
    pattern = re.compile("|".join(parts), re.IGNORECASE)

    stats = {}

    def repl(m):
        kind = m.lastgroup
        text = m.group()
        low = text.lower()
        if kind == "street":
            out = street_map[low]
        elif kind == "acctnum":
            out = acctnum_map[low]
        elif kind == "llcprop":
            out = llcprop_map[low]
        elif kind == "name":
            out = name_map[low]
        elif kind == "town":
            out = town_map[low]
        elif kind == "propcode":
            out = prop_map[low]
        elif kind == "entcode":
            out = ent_map[low]
        elif kind == "last4":
            out = last4_map.get(text, text)
        elif kind == "gluedpair":
            pair = find_glued_pair(low)
            if pair:
                out = single_names[pair[0]] + single_names[pair[1]]
            else:
                out, kind = text, "gluedpair_skip"
        elif kind == "run":
            rr = random.Random(f"{seed}|run|{text}")
            out = "".join(str(rr.randrange(10)) for _ in text)
        elif kind == "alcode":
            if is_code(text):
                out = fake_alnum(seed, text)
            else:
                out, kind = text, "alcode_skip"
        else:
            out = text
        stats[kind] = stats.get(kind, 0) + 1
        return out

    return pattern, repl, stats


def serialize_cell(v) -> str:
    """Faithfully render a cell value for TSV output. Text is returned as-is
    (the caller applies replacement); numbers and dates pass through by value."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else repr(v)
    if isinstance(v, datetime.datetime):
        if v.hour or v.minute or v.second:
            return f"{v.month}/{v.day}/{v.year} {v.hour:02d}:{v.minute:02d}:{v.second:02d}"
        return f"{v.month}/{v.day}/{v.year}"
    if isinstance(v, datetime.date):
        return f"{v.month}/{v.day}/{v.year}"
    return str(v)


def build_filename_replacer9(mapping: dict):
    """Filename-only replacer: swaps in the SAME known real values as
    build_replacer9 (names, towns, account numbers/last-four, property/entity
    codes), but skips its two blanket free-text heuristics - the 5+ digit run
    scrub and the generic mixed alnum 'reference code' scramble. Those exist
    for description/notes text; applied to a filename they can nuke an
    ordinary name that just happens to mix in a few digits (e.g. a workbook
    named '2025MRBusinessOnly' would otherwise be scrambled into gibberish
    even though it contains no real identifier at all).

    The last-four boundary is relaxed to digit-only (not alnum), so a real
    account number glued directly onto letters (e.g. 'Chase7352_Activity...')
    is still caught - that gluing is exactly how export tools name these
    files, and is the actual leak this exists to close."""
    street_map = {r.lower(): f for r, f in mapping["streets"].items()}
    name_map = dict(mapping["names"])
    town_map = dict(mapping["towns"])
    prop_map = dict(mapping["property_codes"])
    ent_map = {k.lower(): v for k, v in mapping["entity_codes"].items()}
    last4_map = dict(mapping["last_four"])
    acctnum_map = {}
    for a in mapping["accounts"]:
        for form in a["real_forms"]:
            if any(not c.isdigit() for c in form):
                acctnum_map[form.lower()] = a["fake_number"]

    parts = []

    def grp(gname, literals, pre=WORDISH_PRE, post=WORDISH_POST):
        body = _alt(literals)
        if body:
            parts.append(rf"(?P<{gname}>{pre}{body}{post})")
    grp("street", list(street_map))
    grp("acctnum", list(acctnum_map))
    grp("name", list(name_map))
    grp("town", list(town_map))
    grp("propcode", list(prop_map), pre=r"(?<![A-Za-z0-9])", post=r"(?![A-Za-z0-9])")
    grp("entcode", list(ent_map), pre=r"(?<![A-Za-z0-9])", post=r"(?![A-Za-z0-9])")
    if last4_map:
        l4 = "(?:" + "|".join(re.escape(k) for k in sorted(last4_map)) + ")"
        parts.append(rf"(?P<last4>(?<!\d){l4}(?!\d))")
    if not parts:
        return None, None

    pattern = re.compile("|".join(parts), re.IGNORECASE)
    maps = {"street": street_map, "acctnum": acctnum_map, "name": name_map,
            "town": town_map, "propcode": prop_map, "entcode": ent_map,
            "last4": last4_map}

    def repl(m):
        kind = m.lastgroup
        text = m.group()
        return maps[kind].get(text.lower(), text) if kind != "last4" \
            else maps[kind].get(text, text)

    return pattern, repl


def anonymize_filename_stem(stem: str, mapping: dict) -> str:
    """Anonymize a file NAME using the SAME known real values as cell text
    (see build_filename_replacer9) - export filenames sometimes carry the
    real account number (e.g. 'Chase7352_Activity_20260725'), so the output
    file must not leak it even though the file's own content is already
    clean. Then keep it a valid Windows filename."""
    pattern, repl = build_filename_replacer9(mapping)
    out = pattern.sub(repl, stem) if pattern else stem
    return re.sub(r'[<>:"/\\|?*]', "_", out)


def apply_workbook9(wb_path: Path, mapping: dict, out_dir: Path, amount_factor: float = 1.0):
    """Anonymize every visible data tab and write ONE plain .xlsx workbook (one
    sheet per tab, no styling). Text cells are replaced; numbers/dates are kept
    as native Excel types so amounts stay numeric and Excel parses it cleanly.
    If amount_factor != 1.0, every 'Amount' cell is also scaled by that
    constant factor (rounded to the cent) - an OPT-IN departure from the
    default 'books tie to the cent' rule, for extra obfuscation on request."""
    pattern, repl, stats = build_replacer9(mapping)
    addr_map = mapping.get("addresses", {})
    # Some sheets (e.g. RulesN) store a bare account number ('585') as a real
    # Excel INTEGER rather than text, unlike the 'xxxx0585' masked form which
    # is always text. Numeric cells skip the text regex entirely, so those
    # bare numbers were passing through un-anonymized. Look them up by the
    # same account key used everywhere else and substitute the fake bare
    # number, keeping the cell numeric (matches the source's own formatting).
    key_to_fake4 = {a["key"]: a["fake_last4"] for a in mapping.get("accounts", [])}
    # Word-based account identifiers ('cash', 'PP' - no real bank number, so
    # acct_key() fell back to the raw text) are deliberately excluded from
    # build_replacer9's global blanket match (see there for why: 'cash' is
    # also an ordinary word elsewhere, e.g. 'Blue Cash Everyday'). They're
    # still real account identifiers though, so substitute them here, scoped
    # to ONLY the Account/Account # cells of their own row.
    word_acctnum_map = {}
    for a in mapping.get("accounts", []):
        for form in a["real_forms"]:
            if not any(c.isdigit() for c in form):
                word_acctnum_map[form.lower()] = a["fake_number"]
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)                      # start empty; add tabs below
    written = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        it = ws.iter_rows(values_only=True)
        row1 = list(next(it, ()) or [])
        width = used_width(row1)
        if width == 0:
            continue
        header = [_clean(v).lower() for v in row1[:width]]
        addr_i = header.index("address") if "address" in header else -1
        acctnum_i = header.index("account #") if "account #" in header else -1
        acct_i = header.index("account") if "account" in header else -1
        inst_i = header.index("institution") if "institution" in header else -1
        amt_i = header.index("amount") if "amount" in header else -1
        # Excel sheet titles: max 31 chars, none of  : \ / ? * [ ]
        title = re.sub(r"[:\\/?*\[\]]", "_", ws.title)[:31] or "Sheet"
        out_ws = out_wb.create_sheet(title)
        out_ws.append(list(row1[:width]))
        n = 0
        for r in it:
            cells = list(r)[:width]
            if not any(c is not None and str(c).strip() for c in cells):
                continue
            row = []
            for ci, v in enumerate(cells):
                if ci == acctnum_i and isinstance(v, (int, float)) and not isinstance(v, bool):
                    fake4 = key_to_fake4.get(acct_key(str(v)))
                    row.append(int(fake4) if fake4 is not None else v)
                elif (ci == amt_i and amount_factor != 1.0
                      and isinstance(v, (int, float)) and not isinstance(v, bool)):
                    row.append(round(v * amount_factor, 2))
                elif not isinstance(v, str):
                    row.append(v)                    # int/float/datetime/None: native
                elif ci in (acct_i, acctnum_i) and v.strip().lower() in word_acctnum_map:
                    # Word-based account identifier ('cash', 'PP') - substitute
                    # directly, scoped to just this cell (see word_acctnum_map).
                    row.append(word_acctnum_map[v.strip().lower()])
                elif ci == inst_i:
                    # Institutions (e.g. 'Chase') are left as-is by design.
                    # Without this carve-out, an account whose Account #/Account
                    # is a plain word rather than a real number (e.g. a 'cash'
                    # pseudo-account with Account = Account # = Institution =
                    # 'cash') would have its Institution cell corrupted too,
                    # since the same literal text is also a harvested account
                    # identifier and gets swapped everywhere it appears.
                    row.append(v)
                elif ci == addr_i:
                    # Address cell: use the address mapping (keeps state/ZIP);
                    # fall back to the regex if this exact address wasn't harvested.
                    row.append(addr_map.get(v) or pattern.sub(repl, v))
                else:
                    row.append(pattern.sub(repl, v))
            out_ws.append(row)
            n += 1
        written.append((title, n))
    wb.close()
    safe_stem = anonymize_filename_stem(wb_path.stem, mapping)
    out_path = out_dir / f"{safe_stem}_anon.xlsx"
    out_wb.save(out_path)
    out_wb.close()
    return written, out_path, stats


def cmd_apply9(args) -> int:
    src: Path = args.workbook
    if not src.exists() or not (args.config and args.config.exists()):
        print("ERROR: workbook and --config are required and must exist.", file=sys.stderr)
        return 2
    print("Draft 9 - increment 4: case-insensitive replacement -> output/<seed>/")
    print(f"  workbook: {src}\n  seed:     {args.seed}")

    harvest = harvest9(src)
    config = load_config(args.config)
    pools = load_pools()
    mapping = build_mapping9(harvest, config, pools, args.seed)
    map_out = MAPPINGS9_DIR / str(args.seed) / "mapping.json"
    map_out.parent.mkdir(parents=True, exist_ok=True)
    with open(map_out, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False, sort_keys=True)
        f.write("\n")

    out_dir = (args.out or (src.parent / "output")) / str(args.seed)
    written, wb_out, stats = apply_workbook9(src, mapping, out_dir)

    print(f"\n  output folder: {out_dir}")
    print(f"    {wb_out.name}  ({len(written)} tabs)")
    for title, n in written:
        print(f"      - {title}: {n} data rows")
    print("  replacements by rule:")
    for k in ("street", "acctnum", "name", "town", "propcode", "entcode", "last4", "run", "alcode"):
        if k in stats:
            print(f"    {k:10s}: {stats[k]}")
    print(f"  mapping: {map_out}  (re-identification key - keep private)")
    print("\nRESULT: PASS - data tabs anonymized to output/<seed>/ (bank CSVs handled by the folder runner).")
    return 0


def _workbook_text(path: Path) -> str:
    """All text-cell content of the visible data tabs, joined — for the scan
    self-test (the blacklist roots should appear in the source)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    chunks = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        it = ws.iter_rows(values_only=True)
        width = used_width(list(next(it, ()) or []))
        for r in it:
            for v in list(r)[:width]:
                if isinstance(v, str) and v.strip():
                    chunks.append(v)
    wb.close()
    return "\n".join(chunks)


def scan_blacklist9(out_dir: Path, terms):
    """Case-insensitive, whitespace-trimmed WHOLE-WORD scan of every output file
    (word boundaries avoid false positives like 'Amira' for the root 'Amir').
    Returns leaks [(file, term, context)]."""
    pats = [(t.strip(), re.compile(WORDISH_PRE + re.escape(t.strip()) + WORDISH_POST, re.IGNORECASE))
            for t in terms if t.strip()]
    leaks = []
    files = sorted(list(out_dir.glob("*.txt")) + list(out_dir.glob("*.csv"))
                   + list(out_dir.glob("*.xlsx")))
    for f in files:
        text = _workbook_text(f) if f.suffix.lower() == ".xlsx" \
            else f.read_text(encoding="utf-8-sig", errors="replace")
        for tt, rx in pats:
            m = rx.search(text)
            if m:
                i = m.start()
                ctx = text[max(0, i - 25):i + len(tt) + 25].replace("\n", " ").replace("\t", " ")
                leaks.append((f.name, tt, ctx))
    return leaks


def cmd_scan9(args) -> int:
    src: Path = args.workbook
    if not (args.config and args.config.exists()):
        print("ERROR: --config is required.", file=sys.stderr)
        return 2
    out_dir = (args.out or (src.parent / "output")) / str(args.seed)
    if not out_dir.exists():
        print(f"ERROR: output folder not found: {out_dir}\n  Run apply9 for this seed first.", file=sys.stderr)
        return 2
    print("Draft 9 - increment 5: blacklist leakage scan (case-insensitive, whole-word)")
    print(f"  outputs:   {out_dir}")

    config = load_config(args.config)
    terms = config["Blacklist"]["distinct"]

    # Self-test: the roots should be present in the SOURCE (or the scan is moot).
    src_text = _workbook_text(src).lower() if src.exists() else ""
    found_src = sum(1 for t in terms if t.strip() and t.strip().lower() in src_text)
    print(f"  blacklist terms: {len([t for t in terms if t.strip()])}   "
          f"present in source (self-test): {found_src}")

    leaks = scan_blacklist9(out_dir, terms)
    print(f"\n  leaks in output (must be 0): {len(leaks)}")
    for fn, t, ctx in leaks[:40]:
        print(f"    [{fn}] {t!r} in ...{ctx}...")
    if len(leaks) > 40:
        print(f"    ... and {len(leaks) - 40} more")

    ok = len(leaks) == 0
    print("\nRESULT:", "PASS - no blacklist root survives in any output."
          if ok else "FAIL - blacklist leakage (see above).")
    return 0 if ok else 1


def cmd_wbinspect(args) -> int:
    src: Path = args.workbook
    if not src.exists():
        print(f"ERROR: workbook not found: {src}", file=sys.stderr)
        return 2
    print("Draft 9 - increment 1: load workbook read-only; list data tabs; clamp columns")
    print(f"  workbook: {src}")

    tabs = inspect_workbook9(src)
    print(f"\nSheets ({len(tabs)}):")
    for t in tabs:
        mark = "PROCESS" if t["processed"] else " skip  "
        clamp = (f"cols {t['reported_cols']}->{t['used_cols']}"
                 if t["reported_cols"] != t["used_cols"] else f"cols {t['used_cols']}")
        print(f"  [{mark}] {t['title']!r:28} state={t['state']:10} rows~{t['reported_rows']:<6} {clamp}")
        if t["processed"]:
            print(f"            header: {t['header']}")

    if args.config and args.config.exists():
        cfg = load_config(args.config)
        print("\nConfig CSV (Names / Towns / Blacklist):")
        for name in CONFIG_COLUMNS:
            print(f"  {name}: {len(cfg[name]['distinct'])} distinct")

    processed = [t for t in tabs if t["processed"]]
    print(f"\nRESULT: PASS - workbook opened read-only; {len(processed)} data tab(s) to process; "
          "phantom columns clamped. (No mapping/replacement yet.)")
    return 0


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="Anonymizer (incremental build).")
    p.add_argument("--version", action="version", version=f"anonymizer v{VERSION}")
    sub = p.add_subparsers(dest="command", required=True)

    rt = sub.add_parser("roundtrip", help="Increment 1: no-op round-trip gate.")
    rt.add_argument("input", type=Path, help="Source workbook (.xlsx/.xlsm)")
    rt.add_argument("-o", "--out", type=Path, default=None,
                    help="Output path (default: ./output/<name>_roundtrip<ext>)")
    rt.set_defaults(func=cmd_roundtrip)

    ins = sub.add_parser("inspect", help="Increment 2: load + report columns/counts.")
    ins.add_argument("input", type=Path, help="Source workbook (.xlsx/.xlsm)")
    ins.add_argument("-c", "--config", type=Path, required=True,
                     help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    ins.add_argument("-s", "--sheets", nargs="+", default=None,
                     help=f"Sheets to inspect (default: {DEFAULT_SHEETS} if present)")
    ins.set_defaults(func=cmd_inspect)

    ext = sub.add_parser("extract", help="Increment 3: extract identifiers (detection only).")
    ext.add_argument("input", type=Path, help="Source workbook (.xlsx/.xlsm)")
    ext.add_argument("-c", "--config", type=Path, required=True,
                     help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    ext.add_argument("-s", "--sheets", nargs="+", default=None,
                     help=f"Sheets to pool (default: {DEFAULT_SHEETS} if present)")
    ext.set_defaults(func=cmd_extract)

    mp = sub.add_parser("map", help="Increment 4: build + write the deterministic mapping.")
    mp.add_argument("input", type=Path, help="Source workbook (.xlsx/.xlsm)")
    mp.add_argument("-c", "--config", type=Path, required=True,
                    help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    mp.add_argument("--seed", default=None, help="Seed (optional; auto-generated if omitted, recorded in the mapping)")
    mp.add_argument("--map-out", type=Path, default=None,
                    help="Mapping table path (default: <code folder>/Anon mapping/mapping.<seed>.json)")
    mp.add_argument("-s", "--sheets", nargs="+", default=None,
                    help=f"Sheets to pool (default: {DEFAULT_SHEETS} if present)")
    mp.set_defaults(func=cmd_map)

    ap = sub.add_parser("apply", help="Increment 5: single-pass replacement over both sheets.")
    ap.add_argument("input", type=Path, help="Source workbook (.xlsx/.xlsm)")
    ap.add_argument("-c", "--config", type=Path, required=True,
                    help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    ap.add_argument("--seed", default=None, help="Seed (optional; auto-generated if omitted)")
    ap.add_argument("-o", "--out", type=Path, default=None,
                    help="Anonymized workbook path (default: ./output/<name>_anon.<seed><ext>)")
    ap.add_argument("-s", "--sheets", nargs="+", default=None,
                    help=f"Sheets to process (default: {DEFAULT_SHEETS} if present)")
    ap.set_defaults(func=cmd_apply)

    sc = sub.add_parser("scan", help="Increment 6: blacklist / real-side leakage scan.")
    sc.add_argument("input", type=Path, help="Original source workbook (.xlsx/.xlsm)")
    sc.add_argument("-c", "--config", type=Path, required=True,
                    help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    sc.add_argument("--anon", type=Path, default=None,
                    help="Anonymized workbook to scan (default: ./<input dir>/output/<name>_anon<ext>)")
    sc.add_argument("-s", "--sheets", nargs="+", default=None,
                    help=f"Sheets to scan (default: {DEFAULT_SHEETS} if present)")
    sc.set_defaults(func=cmd_scan)

    hc = sub.add_parser("hitcount", help="Increment 7: account/transfer hit-count check.")
    hc.add_argument("input", type=Path, help="Original source workbook (.xlsx/.xlsm)")
    hc.add_argument("-c", "--config", type=Path, required=True,
                    help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    hc.add_argument("--seed", required=True, help="Seed (must match the apply run)")
    hc.add_argument("--anon", type=Path, default=None,
                    help="Anonymized workbook (default: ./<input dir>/output/<name>_anon<ext>)")
    hc.add_argument("-s", "--sheets", nargs="+", default=None,
                    help=f"Sheets (default: {DEFAULT_SHEETS} if present)")
    hc.set_defaults(func=cmd_hitcount)

    rn = sub.add_parser("run", help="Increment 8: one-command anonymize + full report.")
    rn.add_argument("input", type=Path, help="Source workbook (.xlsx/.xlsm)")
    rn.add_argument("-c", "--config", type=Path, required=True,
                    help="Combined config file (.csv/.xlsx) with Names/Towns/Blacklist")
    rn.add_argument("--seed", default=None, help="Seed (optional; auto-generated if omitted, recorded in the mapping)")
    rn.add_argument("-o", "--out", type=Path, default=None,
                    help="Anonymized workbook path (default: ./<input dir>/output/<name>_anon.<seed><ext>)")
    rn.add_argument("--map-out", type=Path, default=None,
                    help="Mapping table path (default: <code folder>/Anon mapping/mapping.<seed>.json)")
    rn.add_argument("-s", "--sheets", nargs="+", default=None,
                    help=f"Sheets to process (default: {DEFAULT_SHEETS} if present)")
    rn.set_defaults(func=cmd_run)

    fc = sub.add_parser("fromconfig", help="Run everything from a settings file (no CLI args).")
    fc.add_argument("file", type=Path, nargs="?", default=None,
                    help="Settings file (default: run.config.txt next to the code)")
    fc.add_argument("--banks-only", action="store_true",
                    help="Anonymize only the bank CSVs listed in the settings file, reusing the mapping.")
    fc.set_defaults(func=cmd_fromconfig)

    wi = sub.add_parser("wbinspect", help="Draft 9 increment 1: load workbook read-only, list/clamp data tabs.")
    wi.add_argument("workbook", type=Path, help="Master Excel workbook (.xlsx/.xlsm)")
    wi.add_argument("-c", "--config", type=Path, default=None,
                    help="Config CSV (Names/Towns/Blacklist) to summarize")
    wi.set_defaults(func=cmd_wbinspect)

    hv = sub.add_parser("harvest9", help="Draft 9 increment 2: harvest structural identifiers + inventory.")
    hv.add_argument("workbook", type=Path, help="Master Excel workbook (.xlsx/.xlsm)")
    hv.add_argument("-c", "--config", type=Path, default=None,
                    help="Config CSV (Names/Towns/Blacklist) to summarize")
    hv.set_defaults(func=cmd_harvest9)

    m9 = sub.add_parser("map9", help="Draft 9 increment 3: build the deterministic bijective mapping.")
    m9.add_argument("workbook", type=Path, help="Master Excel workbook (.xlsx/.xlsm)")
    m9.add_argument("-c", "--config", type=Path, required=True,
                    help="Config CSV (Names/Towns/Blacklist)")
    m9.add_argument("--seed", required=True, help="Seed (names the mapping folder)")
    m9.set_defaults(func=cmd_map9)

    a9 = sub.add_parser("apply9", help="Draft 9 increment 4: replacement pass -> output/<seed>/.")
    a9.add_argument("workbook", type=Path, help="Master Excel workbook (.xlsx/.xlsm)")
    a9.add_argument("-c", "--config", type=Path, required=True,
                    help="Config CSV (Names/Towns/Blacklist)")
    a9.add_argument("--seed", required=True, help="Seed (names the output subfolder + mapping)")
    a9.add_argument("-o", "--out", type=Path, default=None,
                    help="Output base folder (default: <workbook dir>/output)")
    a9.set_defaults(func=cmd_apply9)

    s9 = sub.add_parser("scan9", help="Draft 9 increment 5: blacklist leakage scan over output/<seed>/.")
    s9.add_argument("workbook", type=Path, help="Master Excel workbook (for the self-test)")
    s9.add_argument("-c", "--config", type=Path, required=True,
                    help="Config CSV (Names/Towns/Blacklist)")
    s9.add_argument("--seed", required=True, help="Seed (locates output/<seed>/)")
    s9.add_argument("-o", "--out", type=Path, default=None,
                    help="Output base folder (default: <workbook dir>/output)")
    s9.set_defaults(func=cmd_scan9)

    r9 = sub.add_parser("run9",
                        help="Draft 9 increment 7: one command -> workbook + bank CSVs + checks + report.")
    r9.add_argument("workbook", type=Path, help="Master Excel workbook (.xlsx/.xlsm)")
    r9.add_argument("-c", "--config", type=Path, required=True,
                    help="Config CSV (Names/Towns/Blacklist)")
    r9.add_argument("--seed", required=True, help="Seed (names the output subfolder + mapping)")
    r9.add_argument("--banks", nargs="*", default=[],
                    help="Zero or more bank statement CSV paths to anonymize with the same mapping")
    r9.add_argument("-o", "--out", type=Path, default=None,
                    help="Output base folder (default: <workbook dir>/output)")
    r9.add_argument("--amount-factor", type=float, default=1.0,
                    help="Scale every dollar amount by this factor, 0.80-1.20 (default 1.0 = off)")
    r9.set_defaults(func=cmd_run9)

    fc9 = sub.add_parser("fromconfig9",
                         help="Draft 9 increment 7: run everything from run9.config.txt (double-click).")
    fc9.add_argument("-f", "--file", type=Path, default=None,
                     help="Settings file (default: run9.config.txt next to anonymizer.py)")
    fc9.set_defaults(func=cmd_fromconfig9)

    args = p.parse_args(argv)
    print(f"anonymizer v{VERSION}")
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
